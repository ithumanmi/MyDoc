"""
Gộp toàn bộ sao kê Vietcombank trong data/expenses/raw
thành 1 bảng giao dịch để phân tích.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
DATA = REPO_ROOT / "data" / "expenses"
RAW_DIR = DATA / "raw"
WORKING = DATA / "working"

FOLDER = RAW_DIR
OUTPUT_XLSX = WORKING / "bank_transactions_all.xlsx"
OUTPUT_CSV = WORKING / "bank_transactions_all.csv"

HEADER_MARKERS = ("STT", "Debit", "ghi nợ")


def _norm(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return " ".join(str(value).replace("\n", " ").split())


def _parse_money(value: object) -> float | None:
    text = _norm(value)
    if not text or text in {"-", "–"}:
        return None
    # "7,215,820 VND" / "45,000"
    text = re.sub(r"[^\d,.\-]", "", text).replace(",", "")
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date_doc(value: object) -> tuple[pd.Timestamp | None, str]:
    """
    Cột dạng: '01/08/2026 5378 - 39188'
    → ngày + phần chứng từ còn lại.
    """
    text = _norm(value)
    if not text:
        return None, ""
    m = re.match(r"^(\d{1,2}/\d{1,2}/\d{4})\s*(.*)$", text)
    if not m:
        # thử datetime sẵn
        dt = pd.to_datetime(text, dayfirst=True, errors="coerce")
        return (None if pd.isna(dt) else pd.Timestamp(dt)), text
    date = pd.to_datetime(m.group(1), dayfirst=True, errors="coerce")
    doc = m.group(2).strip(" -")
    return (None if pd.isna(date) else pd.Timestamp(date)), doc


def _extract_meta(df: pd.DataFrame) -> dict[str, str]:
    meta = {
        "Chủ tài khoản": "",
        "Số tài khoản": "",
        "Kỳ từ": "",
        "Kỳ đến": "",
        "Loại tiền": "",
    }
    for i in range(min(15, len(df))):
        cells = [_norm(v) for v in df.iloc[i].tolist()]
        row = " | ".join(cells)
        if "Account name" in row or "Chủ tài khoản" in row:
            # value thường ở cột index 3
            if len(cells) > 3 and cells[3]:
                meta["Chủ tài khoản"] = cells[3]
        if "Account number" in row or "Số tài khoản" in row:
            if len(cells) > 3 and cells[3]:
                meta["Số tài khoản"] = cells[3]
        if "From:" in row or "Từ/" in row:
            # ['', 'Từ/ From:', '01/08/2026', 'Đến/ To:', '15/08/2026', '']
            non_empty = [c for c in cells if c]
            for j, c in enumerate(non_empty):
                if "From" in c or "Từ/" in c:
                    if j + 1 < len(non_empty):
                        meta["Kỳ từ"] = non_empty[j + 1]
                if "To" in c or "Đến/" in c:
                    if j + 1 < len(non_empty):
                        meta["Kỳ đến"] = non_empty[j + 1]
        if "Currency" in row or "Loại tiền" in row:
            if len(cells) > 3 and cells[3]:
                meta["Loại tiền"] = cells[3]
    return meta


def _find_header_row(df: pd.DataFrame) -> int | None:
    for i in range(len(df)):
        row = " ".join(_norm(v) for v in df.iloc[i].tolist())
        if "STT" in row and ("Debit" in row or "ghi nợ" in row.lower()):
            return i
    return None


def load_statement(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, header=None)
    meta = _extract_meta(df)
    header_idx = _find_header_row(df)
    if header_idx is None:
        raise ValueError(f"{path.name}: không tìm thấy dòng header giao dịch")

    rows: list[dict] = []
    for i in range(header_idx + 1, len(df)):
        stt_raw = df.iloc[i, 1] if df.shape[1] > 1 else None
        stt_text = _norm(stt_raw)
        if not stt_text.isdigit():
            # hết khối giao dịch (Total / Closing balance...)
            break

        date_doc = df.iloc[i, 2] if df.shape[1] > 2 else None
        debit = _parse_money(df.iloc[i, 3] if df.shape[1] > 3 else None)
        credit = _parse_money(df.iloc[i, 4] if df.shape[1] > 4 else None)
        balance = _parse_money(df.iloc[i, 5] if df.shape[1] > 5 else None)
        detail = _norm(df.iloc[i, 6] if df.shape[1] > 6 else None)
        txn_date, doc_no = _parse_date_doc(date_doc)

        if debit is not None and debit != 0:
            amount = -abs(debit)
            direction = "Chi"
        elif credit is not None and credit != 0:
            amount = abs(credit)
            direction = "Thu"
        else:
            amount = 0.0
            direction = ""

        rows.append(
            {
                "STT_file": int(stt_text),
                "Ngày giao dịch": txn_date,
                "Số chứng từ": doc_no,
                "Số tiền ghi nợ": debit,
                "Số tiền ghi có": credit,
                "Số tiền": amount,
                "Loại giao dịch": direction,
                "Số dư": balance,
                "Nội dung": detail,
                "Chủ tài khoản": meta["Chủ tài khoản"],
                "Số tài khoản": meta["Số tài khoản"],
                "Loại tiền": meta["Loại tiền"] or "VND",
                "Kỳ từ": meta["Kỳ từ"],
                "Kỳ đến": meta["Kỳ đến"],
                "Nguồn file": path.name,
            }
        )

    return pd.DataFrame(rows)


def merge_all(folder: Path = FOLDER) -> pd.DataFrame:
    files = sorted(
        f
        for f in folder.glob("*.xlsx")
        if not f.name.startswith("~$")
    )
    if not files:
        raise FileNotFoundError(f"Không có file Excel trong: {folder}")

    frames = []
    for f in files:
        part = load_statement(f)
        print(f"  - {f.name}: {len(part)} giao dịch")
        frames.append(part)

    all_df = pd.concat(frames, ignore_index=True)

    # Dedup nếu file kỳ chồng nhau
    before = len(all_df)
    all_df = all_df.drop_duplicates(
        subset=["Ngày giao dịch", "Số chứng từ", "Số tiền", "Nội dung", "Số tài khoản"],
        keep="first",
    ).reset_index(drop=True)
    removed = before - len(all_df)

    all_df = all_df.sort_values(
        ["Ngày giao dịch", "STT_file", "Nguồn file"],
        ascending=[True, True, True],
        na_position="last",
    ).reset_index(drop=True)
    all_df.insert(0, "STT", range(1, len(all_df) + 1))

    # Helper thời gian để phân tích
    dates = pd.to_datetime(all_df["Ngày giao dịch"], errors="coerce")
    all_df["Năm"] = dates.dt.year
    all_df["Tháng"] = dates.dt.strftime("%Y-%m")
    all_df["Tuần"] = dates.dt.strftime("%G-W%V")
    all_df["Ngày"] = dates.dt.normalize()

    print(f"Tổng sau gộp: {before} | Sau bỏ trùng: {len(all_df)} (bỏ {removed})")
    return all_df


def _write_excel(df: pd.DataFrame, path: Path) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Giao dịch")
        ws = writer.sheets["Giao dịch"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        money_cols = {
            "Số tiền ghi nợ",
            "Số tiền ghi có",
            "Số tiền",
            "Số dư",
        }
        for idx, col in enumerate(df.columns, start=1):
            letter = get_column_letter(idx)
            if col in money_cols:
                for row in range(2, len(df) + 2):
                    ws[f"{letter}{row}"].number_format = "#,##0"
                ws.column_dimensions[letter].width = 14
            elif col == "Nội dung":
                ws.column_dimensions[letter].width = 55
            elif col in {"Ngày giao dịch", "Ngày"}:
                for row in range(2, len(df) + 2):
                    ws[f"{letter}{row}"].number_format = "dd/mm/yyyy"
                ws.column_dimensions[letter].width = 12
            else:
                sample = df[col].head(30).astype(str).tolist()
                width = max([len(str(col))] + [len(s) for s in sample], default=10)
                ws.column_dimensions[letter].width = min(width + 2, 40)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"


def main() -> None:
    print(f"Đọc folder: {FOLDER}")
    df = merge_all(FOLDER)

    chi = df[df["Loại giao dịch"] == "Chi"]["Số tiền"].sum()
    thu = df[df["Loại giao dịch"] == "Thu"]["Số tiền"].sum()
    print(f"Chi: {chi:,.0f} | Thu: {thu:,.0f} | Net: {chi + thu:,.0f}")
    print(f"Kỳ dữ liệu: {df['Ngày giao dịch'].min()} → {df['Ngày giao dịch'].max()}")

    WORKING.mkdir(parents=True, exist_ok=True)
    _write_excel(df, OUTPUT_XLSX)
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"Đã ghi: {OUTPUT_XLSX}")
    print(f"Đã ghi: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
