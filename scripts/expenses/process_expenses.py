"""
Convert CostData.xlsx → chi_tieu_bi_tru.xlsx:
  - Lọc giao dịch bị trừ (Số tiền < 0)
  - Thêm Ngày / Tuần / Tháng / Năm, Category / SubCategory theo rule
  - Giữ Category/SubCategory đã fill tay từ file chi_tieu cũ (khớp giao dịch)
  - Sheet Tổng hợp + Dashboard tương tác
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

COL_NAME = "Tên định danh"
COL_SUBCATEGORY = "SubCategory"
COL_CATEGORY = "Category"
COL_AMOUNT = "Số tiền"
COL_TIME = "Thời gian"
COL_TYPE = "Loại giao dịch"
COL_DATE = "Ngày"
COL_WEEK = "Tuần"
COL_MONTH = "Tháng"
COL_YEAR = "Năm"

DATE_HELPER_COLS = (COL_DATE, COL_WEEK, COL_MONTH, COL_YEAR)

PARENT_CATEGORIES = ("Thiết yếu", "Lựa chọn", "Văn hóa", "Ngoài dự tính")

# Màu nền Excel (hex không có #)
CATEGORY_COLORS: dict[str, str] = {
    "Thiết yếu": "C6EFCE",       # xanh lá nhạt
    "Lựa chọn": "FCE4D6",        # cam nhạt
    "Văn hóa": "DDEBF7",         # xanh dương nhạt
    "Ngoài dự tính": "F8CBAD",  # cam đậm hơn
}

SUBCATEGORY_COLORS: dict[str, str] = {
    "Ăn Uống": "FFF2CC",
    "Trọ": "E2EFDA",
    "Xăng": "F4B183",
    "4G, Nạp tiền thuê bao": "D0E2F3",
    "Di Chuyển": "C9E4F5",
    "Xe": "BDD7EE",
    "Thuốc": "E2D5F1",
    "Gym": "D9EAD3",
    "SkinCare": "FCE4EC",
    "Mua Sắm": "D5A6BD",
    "Đổi tiền": "EA9999",
}

# Rule: (keyword, SubCategory, Category)
# Khớp không phân biệt hoa thường; rule trên được ưu tiên nếu trùng.
CLASSIFY_RULES: list[tuple[str, str, str]] = [
    # Thiết yếu — Ăn Uống (rule cụ thể trước, "cơm" để sau cùng)
    ("bách hóa xanh", "Ăn Uống", "Thiết yếu"),
    ("quan com co tam", "Ăn Uống", "Thiết yếu"),
    ("cơm tấm", "Ăn Uống", "Thiết yếu"),
    ("com tam", "Ăn Uống", "Thiết yếu"),
    ("bánh mì", "Ăn Uống", "Thiết yếu"),
    ("châu hoàng mỹ hảo", "Ăn Uống", "Thiết yếu"),
    ("hoang thi hue", "Ăn Uống", "Thiết yếu"),
    ("hán văn tư", "Ăn Uống", "Thiết yếu"),
    ("tran thi my tien", "Ăn Uống", "Thiết yếu"),
    ("tran thi hien", "Ăn Uống", "Thiết yếu"),
    ("cơm", "Ăn Uống", "Thiết yếu"),
    # Thiết yếu — khác
    ("petrolimex", "Xăng", "Thiết yếu"),
    ("xang dau dau", "Xăng", "Thiết yếu"),
    ("viettel", "4G, Nạp tiền thuê bao", "Thiết yếu"),
    ("bui nguyen tin", "Trọ", "Thiết yếu"),
    ("nguyễn thanh tùng", "Trọ", "Thiết yếu"),
    ("khach phuong trang futa buslines", "Di Chuyển", "Thiết yếu"),
    ("futabuslines", "Di Chuyển", "Thiết yếu"),
    ("futa", "Di Chuyển", "Thiết yếu"),
    ("pharmacity", "Thuốc", "Thiết yếu"),
    ("thuốc", "Thuốc", "Thiết yếu"),
    # Lựa chọn — Ăn Uống
    ("coffee", "Ăn Uống", "Lựa chọn"),
    ("phê la", "Ăn Uống", "Lựa chọn"),
    ("every half", "Ăn Uống", "Lựa chọn"),
    ("ca phe", "Ăn Uống", "Lựa chọn"),
    # Lựa chọn — Mua Sắm
    ("gs25", "Mua Sắm", "Lựa chọn"),
    ("ministop", "Mua Sắm", "Lựa chọn"),
    # Văn hóa
    ("waysstation", "Gym", "Văn hóa"),
    ("guardian", "SkinCare", "Văn hóa"),
    ("tiktok shop seller", "Mua Sắm", "Văn hóa"),
    ("nguyen thanh lam", "Mua Sắm", "Văn hóa"),
    ("circle k", "Mua Sắm", "Văn hóa"),
    ("tạp hóa", "Mua Sắm", "Văn hóa"),
    ("tap hoa", "Mua Sắm", "Văn hóa"),
    ("trinh huong", "Mua Sắm", "Văn hóa"),
    ("wincommerce", "Mua Sắm", "Văn hóa"),
    # Ngoài dự tính
    ("vo duy luan", "Đổi tiền", "Ngoài dự tính"),
]

# Fallback khi không khớp keyword (SubCategory fill tay → Category)
SUBCATEGORY_TO_CATEGORY: dict[str, str] = {
    "Trọ": "Thiết yếu",
    "Xăng": "Thiết yếu",
    "4G, Nạp tiền thuê bao": "Thiết yếu",
    "Di Chuyển": "Thiết yếu",
    "Xe": "Thiết yếu",
    "Thuốc": "Thiết yếu",
    "Gym": "Văn hóa",
    "SkinCare": "Văn hóa",
    "Mua Sắm": "Văn hóa",
    "Đổi tiền": "Ngoài dự tính",
    # Ăn Uống không map ở đây — phụ thuộc keyword (Thiết yếu / Lựa chọn)
}

REPO_ROOT = Path(__file__).resolve().parents[2]
DATA = REPO_ROOT / "data" / "expenses"
WORKING = DATA / "working"

SOURCE_RAW = "CostData.xlsx"
OUTPUT_FILE = "chi_tieu_bi_tru.xlsx"
DATA_SHEET = "Chi tiêu bị trừ"


def _normalize_columns(columns: pd.Index) -> list[str]:
    return [" ".join(str(c).split()) for c in columns]


def _find_column(columns: list[str], *candidates: str) -> str | None:
    lowered = {c.lower(): c for c in columns}
    for cand in candidates:
        key = cand.lower()
        if key in lowered:
            return lowered[key]
        for col in columns:
            if key in col.lower():
                return col
    return None


def _is_blank(value: object) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def classify_by_name(name: object) -> tuple[str, str]:
    """Trả về (SubCategory, Category) theo keyword; không khớp → ('', '')."""
    text = str(name).casefold() if pd.notna(name) else ""
    for keyword, subcategory, category in CLASSIFY_RULES:
        if keyword.casefold() in text:
            return subcategory, category
    return "", ""


def assign_subcategory(name: object) -> str:
    return classify_by_name(name)[0]


def assign_category_from_name(name: object) -> str:
    return classify_by_name(name)[1]


def assign_category_from_sub(subcategory: object) -> str:
    if _is_blank(subcategory):
        return ""
    return SUBCATEGORY_TO_CATEGORY.get(str(subcategory).strip(), "")


def parse_time_series(series: pd.Series) -> pd.Series:
    """Parse cột Thời gian (datetime / chuỗi dd/mm/yyyy)."""
    parsed = pd.to_datetime(series, dayfirst=True, errors="coerce")
    return parsed


def ensure_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Thêm/cập nhật cột Ngày, Tuần, Tháng, Năm ngay sau Thời gian
    để filter dễ trên Excel. Luôn derive lại từ Thời gian.
    """
    out = df.copy()
    time_col = _find_column(out.columns.tolist(), COL_TIME)
    if time_col is None:
        raise ValueError("Thiếu cột Thời gian")

    # Bỏ helper cũ nếu có (tránh trùng khi chạy lại)
    drop_cols = [
        c
        for c in out.columns
        if c in DATE_HELPER_COLS
        or str(c).strip().lower() in {x.lower() for x in DATE_HELPER_COLS}
    ]
    out = out.drop(columns=drop_cols, errors="ignore")

    parsed = parse_time_series(out[time_col])
    helpers = pd.DataFrame(
        {
            COL_DATE: parsed.dt.normalize(),
            # ISO week: 2026-W50 — filter tuần không lẫn năm
            COL_WEEK: parsed.dt.strftime("%G-W%V"),
            COL_MONTH: parsed.dt.strftime("%Y-%m"),
            COL_YEAR: parsed.dt.year.astype("Int64"),
        },
        index=out.index,
    )
    # NaT → trống cho Tuần/Tháng
    helpers.loc[parsed.isna(), COL_WEEK] = pd.NA
    helpers.loc[parsed.isna(), COL_MONTH] = pd.NA

    insert_at = int(out.columns.get_loc(time_col)) + 1
    left = out.iloc[:, :insert_at]
    right = out.iloc[:, insert_at:]
    return pd.concat([left, helpers, right], axis=1)


def ensure_category_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Đảm bảo có cột Category + SubCategory.
    - Nếu còn cột cũ tên Category chứa SubCategory values → rename thành SubCategory
    - Thêm Category nếu chưa có, đặt trước SubCategory
    """
    out = df.copy()
    name_col = _find_column(out.columns.tolist(), COL_NAME)
    sub_col = _find_column(out.columns.tolist(), COL_SUBCATEGORY)
    cat_col = _find_column(out.columns.tolist(), COL_CATEGORY)

    # Trường hợp cũ: chỉ có Category (= SubCategory thực tế)
    if sub_col is None and cat_col is not None:
        # Đổi tên Category → SubCategory
        out = out.rename(columns={cat_col: COL_SUBCATEGORY})
        sub_col = COL_SUBCATEGORY
        cat_col = None

    if sub_col is None:
        # Tạo SubCategory mới sau tên định danh
        insert_at = (
            out.columns.get_loc(name_col) + 1 if name_col is not None else len(out.columns)
        )
        out.insert(insert_at, COL_SUBCATEGORY, "")
        sub_col = COL_SUBCATEGORY

    if cat_col is None:
        # Thêm Category ngay trước SubCategory
        insert_at = out.columns.get_loc(sub_col)
        out.insert(insert_at, COL_CATEGORY, "")
        cat_col = COL_CATEGORY

    # Chuẩn hóa tên cột nếu tìm thấy variant
    rename_map = {}
    if cat_col != COL_CATEGORY:
        rename_map[cat_col] = COL_CATEGORY
    if sub_col != COL_SUBCATEGORY:
        rename_map[sub_col] = COL_SUBCATEGORY
    if rename_map:
        out = out.rename(columns=rename_map)

    # Đảm bảo thứ tự: ... | Category | SubCategory | ...
    cols = list(out.columns)
    for c in (COL_CATEGORY, COL_SUBCATEGORY):
        if c in cols:
            cols.remove(c)
    if name_col and name_col in cols:
        idx = cols.index(name_col) + 1
        cols[idx:idx] = [COL_CATEGORY, COL_SUBCATEGORY]
    else:
        cols.extend([COL_CATEGORY, COL_SUBCATEGORY])
    return out[cols]


def fill_empty_subcategories(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    out = df.copy()
    name_col = _find_column(out.columns.tolist(), COL_NAME)
    if name_col is None:
        raise ValueError("Thiếu cột Tên định danh")

    empty_mask = out[COL_SUBCATEGORY].map(_is_blank)
    suggested = out.loc[empty_mask, name_col].map(assign_subcategory)
    fill_index = suggested.index[suggested.map(lambda v: not _is_blank(v))]
    out.loc[fill_index, COL_SUBCATEGORY] = suggested.loc[fill_index]
    return out, len(fill_index)


def apply_categories(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Gán Category theo keyword tên định danh; fallback SubCategory map.
    Cập nhật lại Category khi rule cho ra giá trị (taxonomy mới).
    Không đụng SubCategory.
    """
    out = df.copy()
    name_col = _find_column(out.columns.tolist(), COL_NAME)
    if name_col is None:
        raise ValueError("Thiếu cột Tên định danh")

    from_name = out[name_col].map(assign_category_from_name)
    from_sub = out[COL_SUBCATEGORY].map(assign_category_from_sub)
    suggested = from_name.where(from_name.map(lambda v: not _is_blank(v)), from_sub)

    fill_mask = suggested.map(lambda v: not _is_blank(v))
    out.loc[fill_mask, COL_CATEGORY] = suggested.loc[fill_mask]
    return out, int(fill_mask.sum())


def load_cost_data(path: Path) -> pd.DataFrame:
    """Đọc CostData.xlsx, lấy cột cần thiết, chỉ giữ tiền bị trừ."""
    df = pd.read_excel(path)
    df.columns = _normalize_columns(df.columns)

    col_time = _find_column(df.columns.tolist(), "Thời gian")
    col_type = _find_column(df.columns.tolist(), "Loại giao dịch")
    col_name = _find_column(
        df.columns.tolist(),
        "Tên định danh Tài khoản nhận",
        "Tên định danh",
    )
    col_amount = _find_column(df.columns.tolist(), "Số Tiền", "Số tiền")

    missing = [
        label
        for label, col in [
            (COL_TIME, col_time),
            (COL_TYPE, col_type),
            (COL_NAME, col_name),
            (COL_AMOUNT, col_amount),
        ]
        if col is None
    ]
    if missing:
        raise ValueError(f"{path.name}: thiếu cột {', '.join(missing)}")

    out = pd.DataFrame(
        {
            COL_TIME: df[col_time],
            COL_TYPE: df[col_type],
            COL_NAME: df[col_name],
            COL_AMOUNT: pd.to_numeric(df[col_amount], errors="coerce"),
        }
    )
    out["Nguồn file"] = path.name
    out = out[out[COL_AMOUNT] < 0].copy().reset_index(drop=True)
    return out


def _row_match_key(df: pd.DataFrame) -> pd.Series:
    """Khóa khớp giao dịch: thời gian + tên + số tiền."""
    parsed = parse_time_series(df[COL_TIME])
    time_part = parsed.dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
    name_part = df[COL_NAME].map(lambda v: "" if _is_blank(v) else str(v).strip().casefold())
    amt_part = pd.to_numeric(df[COL_AMOUNT], errors="coerce").map(
        lambda v: "" if pd.isna(v) else f"{float(v):.0f}"
    )
    return time_part + "|" + name_part + "|" + amt_part


def _row_match_key_loose(df: pd.DataFrame) -> pd.Series:
    """Khóa phụ: tên + số tiền (khi lệch format thời gian)."""
    name_part = df[COL_NAME].map(lambda v: "" if _is_blank(v) else str(v).strip().casefold())
    amt_part = pd.to_numeric(df[COL_AMOUNT], errors="coerce").map(
        lambda v: "" if pd.isna(v) else f"{float(v):.0f}"
    )
    parsed = parse_time_series(df[COL_TIME])
    day_part = parsed.dt.strftime("%Y-%m-%d").fillna("")
    return day_part + "|" + name_part + "|" + amt_part


def load_previous_labels(path: Path) -> pd.DataFrame:
    """Đọc Category/SubCategory đã có từ chi_tieu_bi_tru.xlsx (nếu có)."""
    if not path.exists():
        return pd.DataFrame(
            columns=["_key", "_key_loose", COL_CATEGORY, COL_SUBCATEGORY]
        )
    try:
        df = pd.read_excel(path, sheet_name=DATA_SHEET)
    except ValueError:
        return pd.DataFrame(
            columns=["_key", "_key_loose", COL_CATEGORY, COL_SUBCATEGORY]
        )

    df.columns = _normalize_columns(df.columns)
    mask_total = df.astype(str).apply(
        lambda col: col.str.contains(r"Tổng \(theo filter\)", na=False, regex=True)
    ).any(axis=1)
    if mask_total.any():
        df = df.loc[~mask_total].copy()

    needed = [COL_TIME, COL_NAME, COL_AMOUNT]
    if any(c not in df.columns for c in needed):
        return pd.DataFrame(
            columns=["_key", "_key_loose", COL_CATEGORY, COL_SUBCATEGORY]
        )

    cat_col = _find_column(df.columns.tolist(), COL_CATEGORY) or COL_CATEGORY
    sub_col = _find_column(df.columns.tolist(), COL_SUBCATEGORY) or COL_SUBCATEGORY
    if cat_col not in df.columns:
        df[COL_CATEGORY] = ""
        cat_col = COL_CATEGORY
    if sub_col not in df.columns:
        df[COL_SUBCATEGORY] = ""
        sub_col = COL_SUBCATEGORY

    labels = pd.DataFrame(
        {
            "_key": _row_match_key(df),
            "_key_loose": _row_match_key_loose(df),
            COL_CATEGORY: df[cat_col].map(
                lambda v: "" if _is_blank(v) else str(v).strip()
            ),
            COL_SUBCATEGORY: df[sub_col].map(
                lambda v: "" if _is_blank(v) else str(v).strip()
            ),
        }
    )
    keep = ~labels[COL_CATEGORY].map(_is_blank) | ~labels[COL_SUBCATEGORY].map(_is_blank)
    labels = labels.loc[keep].drop_duplicates("_key", keep="last")
    return labels


def merge_previous_labels(df: pd.DataFrame, labels: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Gắn lại Category/SubCategory cũ; không đè nếu ô mới đã có."""
    out = df.copy()
    if labels.empty:
        return out, 0

    out["_key"] = _row_match_key(out)
    out["_key_loose"] = _row_match_key_loose(out)

    # Ưu tiên khóa chặt, rồi khóa lỏng cho dòng còn trống
    merged = out.merge(
        labels[["_key", COL_CATEGORY, COL_SUBCATEGORY]],
        on="_key",
        how="left",
        suffixes=("", "_prev"),
    )
    restored = 0
    for col in (COL_SUBCATEGORY, COL_CATEGORY):
        prev = f"{col}_prev"
        if prev not in merged.columns:
            continue
        empty = merged[col].map(_is_blank)
        has_prev = ~merged[prev].map(_is_blank)
        take = empty & has_prev
        restored += int(take.sum())
        merged.loc[take, col] = merged.loc[take, prev]
    merged = merged.drop(
        columns=[c for c in merged.columns if c.endswith("_prev")],
        errors="ignore",
    )

    still_empty_sub = merged[COL_SUBCATEGORY].map(_is_blank)
    still_empty_cat = merged[COL_CATEGORY].map(_is_blank)
    if still_empty_sub.any() or still_empty_cat.any():
        loose = labels.drop_duplicates("_key_loose", keep="last")
        merged = merged.merge(
            loose[["_key_loose", COL_CATEGORY, COL_SUBCATEGORY]],
            on="_key_loose",
            how="left",
            suffixes=("", "_loose"),
        )
        for col in (COL_SUBCATEGORY, COL_CATEGORY):
            prev = f"{col}_loose"
            if prev not in merged.columns:
                continue
            empty = merged[col].map(_is_blank)
            has_prev = ~merged[prev].map(_is_blank)
            take = empty & has_prev
            restored += int(take.sum())
            merged.loc[take, col] = merged.loc[take, prev]
        merged = merged.drop(
            columns=[c for c in merged.columns if c.endswith("_loose")],
            errors="ignore",
        )

    merged = merged.drop(columns=["_key", "_key_loose"], errors="ignore")
    return merged, restored


def process_folder(
    folder: Path | None = None,
    raw_name: str = SOURCE_RAW,
    output_name: str = OUTPUT_FILE,
) -> Path:
    folder = folder or WORKING
    folder.mkdir(parents=True, exist_ok=True)
    raw_path = folder / raw_name
    output_path = folder / output_name
    if not raw_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file nguồn: {raw_path}")

    df = load_cost_data(raw_path)
    previous = load_previous_labels(output_path)

    updated = ensure_category_columns(df)
    updated, restored = merge_previous_labels(updated, previous)
    updated = ensure_date_columns(updated)
    updated, new_sub = fill_empty_subcategories(updated)
    updated, new_cat = apply_categories(updated)

    parsed = parse_time_series(updated[COL_TIME])
    updated = updated.assign(_sort=parsed).sort_values("_sort", ascending=False)
    updated = updated.drop(columns=["_sort"]).reset_index(drop=True)

    try:
        _write_excel(updated, output_path)
    except PermissionError:
        alt_path = folder / f"{Path(output_name).stem}_out{Path(output_name).suffix}"
        _write_excel(updated, alt_path)
        print(
            f"Không ghi được {output_path.name} (file đang mở). "
            f"Đã ghi tạm: {alt_path.name}"
        )
        output_path = alt_path

    def counts(series: pd.Series) -> pd.Series:
        return series.map(
            lambda v: "(chưa gán)" if _is_blank(v) else str(v).strip()
        ).value_counts()

    print(f"Nguồn: {raw_path.name}")
    print(f"Giao dịch bị trừ: {len(updated)}")
    print(f"Nhãn cũ giữ lại: {restored}")
    print(f"SubCategory mới điền: {new_sub}")
    print(f"Category gán từ rule: {new_cat}")
    print("Category:")
    for cat, count in counts(updated[COL_CATEGORY]).items():
        print(f"  - {cat}: {count}")
    print("SubCategory:")
    for sub, count in counts(updated[COL_SUBCATEGORY]).items():
        print(f"  - {sub}: {count}")

    amount_col = _find_column(updated.columns.tolist(), COL_AMOUNT, "Số Tiền")
    if amount_col:
        total = pd.to_numeric(updated[amount_col], errors="coerce").sum()
        print(f"Tổng số tiền: {total:,.0f}")
        for title, summary_df in _build_period_summary(updated).items():
            print(f"{title}: {len(summary_df)} nhóm")

    print(f"Đã ghi: {output_path}")

    # Sau khi convert MoMo → bổ sung chi trực tiếp từ bank (cột Note)
    bank_file = folder / "bank_transactions_all.xlsx"
    if bank_file.exists() and output_path.name == OUTPUT_FILE:
        try:
            import sync_bank_gaps

            print("\n--- Đồng bộ khoản thiếu từ ngân hàng ---")
            sync_bank_gaps.main()
        except Exception as exc:  # noqa: BLE001
            print(f"Bỏ qua sync bank gaps: {exc}")

    # Export Power BI ready
    try:
        import export_powerbi

        print("\n--- Export Power BI ---")
        export_powerbi.export_all()
    except Exception as exc:  # noqa: BLE001
        print(f"Bỏ qua export Power BI: {exc}")

    return output_path


def _build_period_summary(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Tổng số tiền theo Ngày / Tuần / Tháng."""
    amount_col = _find_column(df.columns.tolist(), COL_AMOUNT, "Số Tiền")
    if amount_col is None:
        return {}

    amount = pd.to_numeric(df[amount_col], errors="coerce")
    summaries: dict[str, pd.DataFrame] = {}

    specs = [
        (COL_DATE, "Theo ngày", "Ngày"),
        (COL_WEEK, "Theo tuần", "Tuần"),
        (COL_MONTH, "Theo tháng", "Tháng"),
    ]
    for col, title, label in specs:
        if col not in df.columns:
            continue
        grouped = (
            pd.DataFrame({label: df[col], "Tổng số tiền": amount})
            .dropna(subset=[label])
            .groupby(label, dropna=True, sort=False)["Tổng số tiền"]
            .sum()
            .reset_index()
        )
        # Sắp xếp mới → cũ
        if col == COL_DATE:
            grouped = grouped.sort_values(label, ascending=False)
        elif col == COL_WEEK:
            grouped = grouped.sort_values(label, ascending=False)
        else:
            grouped = grouped.sort_values(label, ascending=False)
        grouped = grouped.reset_index(drop=True)
        summaries[title] = grouped
    return summaries


def _write_excel(df: pd.DataFrame, output_path: Path) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    sheet_name = "Chi tiêu bị trừ"
    summary_name = "Tổng hợp"
    amount_col = _find_column(df.columns.tolist(), COL_AMOUNT, "Số Tiền")
    time_col = _find_column(df.columns.tolist(), COL_TIME)
    date_col = _find_column(df.columns.tolist(), COL_DATE)

    fill_cache: dict[str, PatternFill] = {}
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    def fill_for(hex_color: str) -> PatternFill:
        if hex_color not in fill_cache:
            fill_cache[hex_color] = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type="solid",
            )
        return fill_cache[hex_color]

    def apply_value_colors(col_name: str, color_map: dict[str, str]) -> None:
        if col_name not in df.columns:
            return
        idx = int(df.columns.get_loc(col_name)) + 1
        letter = get_column_letter(idx)
        for row in range(2, len(df) + 2):
            cell = ws[f"{letter}{row}"]
            value = cell.value
            if value is None or str(value).strip() == "":
                continue
            key = str(value).strip()
            hex_color = color_map.get(key)
            if hex_color:
                cell.fill = fill_for(hex_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    summaries = _build_period_summary(df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _format_col(col_name: str | None, number_format: str, align: str = "center") -> None:
            if col_name is None or col_name not in df.columns:
                return
            idx = int(df.columns.get_loc(col_name)) + 1
            letter = get_column_letter(idx)
            for row in range(2, len(df) + 2):
                cell = ws[f"{letter}{row}"]
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal=align)

        _format_col(time_col, "dd/mm/yyyy hh:mm:ss", "left")
        _format_col(date_col, "dd/mm/yyyy", "center")
        _format_col(COL_YEAR if COL_YEAR in df.columns else None, "0", "center")

        data_end_row = len(df) + 1
        amount_letter = None
        if amount_col is not None:
            amount_idx = int(df.columns.get_loc(amount_col)) + 1
            amount_letter = get_column_letter(amount_idx)
            for row in range(2, data_end_row + 1):
                cell = ws[f"{amount_letter}{row}"]
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

            # Dòng tổng: SUBTOTAL(109) = sum các dòng đang hiện (sau filter)
            total_row = data_end_row + 1
            label_col = max(1, amount_idx - 1)
            label_cell = ws.cell(total_row, label_col, "Tổng (theo filter)")
            label_cell.font = Font(bold=True)
            label_cell.fill = total_fill
            label_cell.alignment = Alignment(horizontal="right")

            total_cell = ws.cell(
                total_row,
                amount_idx,
                f"=SUBTOTAL(109,{amount_letter}2:{amount_letter}{data_end_row})",
            )
            total_cell.font = Font(bold=True)
            total_cell.fill = total_fill
            total_cell.number_format = '#,##0'
            total_cell.alignment = Alignment(horizontal="right")

        apply_value_colors(COL_CATEGORY, CATEGORY_COLORS)
        apply_value_colors(COL_SUBCATEGORY, SUBCATEGORY_COLORS)

        # Highlight Note bank-only
        if "Note" in df.columns:
            note_idx = int(df.columns.get_loc("Note")) + 1
            note_letter = get_column_letter(note_idx)
            note_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            for row in range(2, data_end_row + 1):
                cell = ws[f"{note_letter}{row}"]
                if cell.value and "ngân hàng" in str(cell.value).lower():
                    cell.fill = note_fill

        # Dropdown cho Category
        if COL_CATEGORY in df.columns and len(df) > 0:
            cat_idx = int(df.columns.get_loc(COL_CATEGORY)) + 1
            cat_letter = get_column_letter(cat_idx)
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(PARENT_CATEGORIES) + '"',
                allow_blank=True,
            )
            dv.error = "Chọn một trong: Thiết yếu, Lựa chọn, Văn hóa, Ngoài dự tính"
            dv.errorTitle = "Category không hợp lệ"
            ws.add_data_validation(dv)
            dv.add(f"{cat_letter}2:{cat_letter}{data_end_row}")

        col_widths = {
            COL_TIME: 20,
            COL_DATE: 12,
            COL_WEEK: 10,
            COL_MONTH: 10,
            COL_YEAR: 8,
            COL_CATEGORY: 12,
            COL_SUBCATEGORY: 18,
            COL_AMOUNT: 14,
        }
        for idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(idx)
            if col_name in col_widths:
                ws.column_dimensions[letter].width = col_widths[col_name]
                continue
            sample = df[col_name].head(50).tolist()
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in sample if pd.notna(v)],
                default=10,
            )
            ws.column_dimensions[letter].width = min(max_len + 2, 45)

        # AutoFilter chỉ phủ data (không gồm dòng tổng)
        last_col = get_column_letter(df.shape[1])
        ws.auto_filter.ref = f"A1:{last_col}{data_end_row}"
        ws.freeze_panes = "A2"

        # Sheet tổng hợp theo ngày / tuần / tháng
        ws_sum = writer.book.create_sheet(summary_name)
        start_col = 1
        for title, summary_df in summaries.items():
            title_cell = ws_sum.cell(1, start_col, title)
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="left")

            headers = list(summary_df.columns)
            for offset, header in enumerate(headers):
                cell = ws_sum.cell(2, start_col + offset, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin

            for r_idx, row in enumerate(summary_df.itertuples(index=False), start=3):
                for c_offset, value in enumerate(row):
                    cell = ws_sum.cell(r_idx, start_col + c_offset, value)
                    cell.border = thin
                    if c_offset == 0 and headers[0] == "Ngày" and value is not None:
                        cell.number_format = "dd/mm/yyyy"
                    if headers[c_offset] == "Tổng số tiền":
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right")

            # Tổng cộng cuối bảng
            end_data_row = 2 + len(summary_df)
            total_label = ws_sum.cell(end_data_row + 1, start_col, "Tổng cộng")
            total_label.font = Font(bold=True)
            total_label.fill = total_fill
            total_label.border = thin
            money_col = start_col + 1
            total_formula = (
                f"=SUM({get_column_letter(money_col)}3:{get_column_letter(money_col)}{end_data_row})"
                if len(summary_df) > 0
                else 0
            )
            total_val = ws_sum.cell(end_data_row + 1, money_col, total_formula)
            total_val.font = Font(bold=True)
            total_val.fill = total_fill
            total_val.number_format = '#,##0'
            total_val.border = thin

            ws_sum.column_dimensions[get_column_letter(start_col)].width = 14
            ws_sum.column_dimensions[get_column_letter(start_col + 1)].width = 16
            start_col += 3

        ws_sum.freeze_panes = "A3"

        _add_dashboard_sheet(writer.book, df)


def _add_dashboard_sheet(workbook, df: pd.DataFrame) -> None:
    """
    Dashboard tương tác:
      - Dropdown filter Tháng / Category / SubCategory
      - KPI + bảng + chart cập nhật theo filter (công thức Excel)
      - Data label trên chart
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    amount_col = _find_column(df.columns.tolist(), COL_AMOUNT, "Số Tiền")
    name_col = _find_column(df.columns.tolist(), COL_NAME)
    if amount_col is None or COL_MONTH not in df.columns:
        return

    cols = list(df.columns)

    def letter(col_name: str) -> str:
        return get_column_letter(cols.index(col_name) + 1)

    data_sheet = DATA_SHEET
    n = len(df)
    data_start, data_end = 2, n + 1  # hàng dữ liệu trên sheet Chi tiêu

    amt_l = letter(amount_col)
    month_l = letter(COL_MONTH)
    week_l = letter(COL_WEEK) if COL_WEEK in cols else None
    cat_l = letter(COL_CATEGORY) if COL_CATEGORY in cols else None
    sub_l = letter(COL_SUBCATEGORY) if COL_SUBCATEGORY in cols else None
    name_l = letter(name_col) if name_col else None

    months = sorted(df[COL_MONTH].dropna().astype(str).unique().tolist())
    categories = [c for c in PARENT_CATEGORIES] + ["(chưa gán)"]
    # SubCategory thực tế trong data
    subs = (
        df[COL_SUBCATEGORY]
        .map(lambda v: "(chưa gán)" if _is_blank(v) else str(v).strip())
        .dropna()
        .unique()
        .tolist()
        if COL_SUBCATEGORY in cols
        else []
    )
    subs = sorted(set(subs))

    latest_month = months[-1] if months else "Tất cả"
    latest_week = (
        sorted(df[COL_WEEK].dropna().astype(str).unique())[-1]
        if COL_WEEK in cols and df[COL_WEEK].notna().any()
        else ""
    )

    if "Dashboard" in workbook.sheetnames:
        del workbook["Dashboard"]
    if "_Lists" in workbook.sheetnames:
        del workbook["_Lists"]

    # Sheet danh sách dropdown (ẩn)
    ws_lists = workbook.create_sheet("_Lists")
    ws_lists["A1"] = "Tháng"
    ws_lists["A2"] = "Tất cả"
    for i, m in enumerate(months, start=3):
        ws_lists[f"A{i}"] = m
    ws_lists["B1"] = "Category"
    ws_lists["B2"] = "Tất cả"
    for i, c in enumerate(categories, start=3):
        ws_lists[f"B{i}"] = c
    ws_lists["C1"] = "SubCategory"
    ws_lists["C2"] = "Tất cả"
    for i, s in enumerate(subs, start=3):
        ws_lists[f"C{i}"] = s
    ws_lists.sheet_state = "hidden"

    month_list_end = 2 + len(months)
    cat_list_end = 2 + len(categories)
    sub_list_end = 2 + len(subs)

    ws = workbook.create_sheet("Dashboard", 0)

    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    title_font = Font(name="Calibri", size=20, bold=True, color="1F4E79")
    kpi_label_font = Font(name="Calibri", size=10, color="666666")
    kpi_value_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    filter_fill = PatternFill("solid", fgColor="FFF2CC")
    filter_font = Font(bold=True, color="C65911")

    PIE_H, BAR_H, LINE_H = 9.0, 10.0, 9.0
    CHART_GAP = 2

    def rows_for_chart(height_cm: float) -> int:
        return max(14, int(height_cm / 0.45) + 2)

    # ---- Title ----
    ws["A1"] = "DASHBOARD CHI TIÊU"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28
    ws["A2"] = (
        "Chọn filter bên dưới → KPI & biểu đồ tự cập nhật "
        "(mở bằng Excel Desktop, bật tính toán công thức)"
    )
    ws["A2"].font = Font(size=9, color="888888")
    ws.merge_cells("A2:F2")

    # ---- Filter panel ----
    # B4=Tháng filter, D4=Category, F4=SubCategory
    ws["A4"] = "FILTER"
    ws["A4"].font = Font(bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill("solid", fgColor="C65911")
    ws["A4"].alignment = Alignment(horizontal="center")

    ws["A5"] = "Tháng"
    ws["B5"] = latest_month
    ws["C5"] = "Category"
    ws["D5"] = "Tất cả"
    ws["E5"] = "SubCategory"
    ws["F5"] = "Tất cả"

    for cell_addr in ("B5", "D5", "F5"):
        cell = ws[cell_addr]
        cell.fill = filter_fill
        cell.font = filter_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for cell_addr in ("A5", "C5", "E5"):
        ws[cell_addr].font = Font(bold=True)
        ws[cell_addr].alignment = Alignment(horizontal="right")

    # Data validation dropdowns
    dv_month = DataValidation(
        type="list",
        formula1=f"=_Lists!$A$2:$A${month_list_end}",
        allow_blank=False,
    )
    dv_cat = DataValidation(
        type="list",
        formula1=f"=_Lists!$B$2:$B${cat_list_end}",
        allow_blank=False,
    )
    dv_sub = DataValidation(
        type="list",
        formula1=f"=_Lists!$C$2:$C${sub_list_end}",
        allow_blank=False,
    )
    ws.add_data_validation(dv_month)
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_sub)
    dv_month.add(ws["B5"])
    dv_cat.add(ws["D5"])
    dv_sub.add(ws["F5"])

    # Named filter cells for formulas
    f_month, f_cat, f_sub = "$B$5", "$D$5", "$F$5"

    def abs_amt_range() -> str:
        return f"ABS('{data_sheet}'!${amt_l}${data_start}:${amt_l}${data_end})"

    def col_range(col_letter: str) -> str:
        return f"'{data_sheet}'!${col_letter}${data_start}:${col_letter}${data_end}"

    def match_filter(filter_cell: str, col_letter: str | None, blank_as: str = "(chưa gán)") -> str:
        """Điều kiện SUMPRODUCT: Tất cả → 1; khớp giá trị; blank Category → (chưa gán)."""
        if col_letter is None:
            return "1"
        rng = col_range(col_letter)
        # (filter="Tất cả") + (range=filter) + (filter=blank_as)*(range="")
        return (
            f"((({filter_cell}=\"Tất cả\")"
            f"+({rng}={filter_cell})"
            f"+(({filter_cell}=\"{blank_as}\")*(({rng}=\"\")+ISBLANK({rng}))))>0)"
        )

    month_cond = match_filter(f_month, month_l)
    cat_cond = match_filter(f_cat, cat_l)
    sub_cond = match_filter(f_sub, sub_l)

    # ---- KPI (công thức) ----
    kpi_fills = ["C6EFCE", "FCE4D6", "DDEBF7"]
    ws.merge_cells("A7:B7")
    ws.merge_cells("A8:B8")
    ws.merge_cells("C7:D7")
    ws.merge_cells("C8:D8")
    ws.merge_cells("E7:F7")
    ws.merge_cells("E8:F8")

    ws["A7"] = "Tổng chi (theo filter)"
    ws["A7"].font = kpi_label_font
    ws["A7"].fill = PatternFill("solid", fgColor=kpi_fills[0])
    ws["A7"].alignment = Alignment(horizontal="center")
    ws["B7"].fill = PatternFill("solid", fgColor=kpi_fills[0])

    # SUMPRODUCT abs amount with filters
    kpi_total_formula = (
        f"=SUMPRODUCT({abs_amt_range()}*{month_cond}*{cat_cond}*{sub_cond})"
    )
    ws["A8"] = kpi_total_formula
    ws["A8"].number_format = '#,##0" đ"'
    ws["A8"].font = kpi_value_font
    ws["A8"].fill = PatternFill("solid", fgColor=kpi_fills[0])
    ws["A8"].alignment = Alignment(horizontal="center")
    ws["B8"].fill = PatternFill("solid", fgColor=kpi_fills[0])

    ws["C7"] = f"Tháng mới nhất ({latest_month})"
    ws["C7"].font = kpi_label_font
    ws["C7"].fill = PatternFill("solid", fgColor=kpi_fills[1])
    ws["C7"].alignment = Alignment(horizontal="center")
    ws["D7"].fill = PatternFill("solid", fgColor=kpi_fills[1])
    month_only = (
        f"=SUMPRODUCT({abs_amt_range()}*"
        f"({col_range(month_l)}=\"{latest_month}\")*{cat_cond}*{sub_cond})"
    )
    ws["C8"] = month_only
    ws["C8"].number_format = '#,##0" đ"'
    ws["C8"].font = kpi_value_font
    ws["C8"].fill = PatternFill("solid", fgColor=kpi_fills[1])
    ws["C8"].alignment = Alignment(horizontal="center")
    ws["D8"].fill = PatternFill("solid", fgColor=kpi_fills[1])

    ws["E7"] = f"Tuần mới nhất ({latest_week})"
    ws["E7"].font = kpi_label_font
    ws["E7"].fill = PatternFill("solid", fgColor=kpi_fills[2])
    ws["E7"].alignment = Alignment(horizontal="center")
    ws["F7"].fill = PatternFill("solid", fgColor=kpi_fills[2])
    if week_l and latest_week:
        week_only = (
            f"=SUMPRODUCT({abs_amt_range()}*"
            f"({col_range(week_l)}=\"{latest_week}\")*{cat_cond}*{sub_cond})"
        )
    else:
        week_only = "0"
    ws["E8"] = week_only
    ws["E8"].number_format = '#,##0" đ"'
    ws["E8"].font = kpi_value_font
    ws["E8"].fill = PatternFill("solid", fgColor=kpi_fills[2])
    ws["E8"].alignment = Alignment(horizontal="center")
    ws["F8"].fill = PatternFill("solid", fgColor=kpi_fills[2])
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 30

    # Số giao dịch theo filter
    ws["A9"] = "Số giao dịch (filter):"
    ws["A9"].font = Font(italic=True, size=9, color="666666")
    count_formula = (
        f"=SUMPRODUCT(({col_range(amt_l)}<>\"\")*{month_cond}*{cat_cond}*{sub_cond})"
    )
    ws["B9"] = count_formula
    ws["B9"].font = Font(bold=True, size=9, color="1F4E79")

    def write_header_row(r: int, headers: list[str]) -> None:
        for i, h in enumerate(headers):
            cell = ws.cell(r, 1 + i, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

    def style_money(cell) -> None:
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin

    row = 11

    # ===== 1) Category table (labels cố định + công thức Chi) =====
    ws.cell(row, 1, "1. Theo Category (theo filter)").font = section_font
    row += 1
    write_header_row(row, ["Category", "Chi"])
    cat_header = row
    row += 1
    cat_first = row
    for i, cat_name in enumerate(categories):
        ws.cell(row, 1, cat_name).border = thin
        if cat_name in CATEGORY_COLORS:
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=CATEGORY_COLORS[cat_name])
        # Chi theo category label, vẫn áp tháng + subcategory filter
        if cat_name == "(chưa gán)":
            cat_match = (
                f"((('{data_sheet}'!${cat_l}${data_start}:${cat_l}${data_end}=\"\")"
                f"+ISBLANK('{data_sheet}'!${cat_l}${data_start}:${cat_l}${data_end}))>0)"
            )
        else:
            cat_match = (
                f"('{data_sheet}'!${cat_l}${data_start}:${cat_l}${data_end}=A{row})"
            )
        formula = (
            f"=SUMPRODUCT({abs_amt_range()}*{month_cond}*{sub_cond}*{cat_match})"
        )
        cell = ws.cell(row, 2, formula)
        style_money(cell)
        row += 1
    cat_last = row - 1
    row += 1

    if cat_last >= cat_first:
        pie = PieChart()
        pie.title = "Chi theo Category"
        labels = Reference(ws, min_col=1, min_row=cat_first, max_row=cat_last)
        data = Reference(ws, min_col=2, min_row=cat_header, max_row=cat_last)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showSerName = False
        pie.width = 16
        pie.height = PIE_H
        ws.add_chart(pie, f"A{row}")
        row += rows_for_chart(PIE_H) + CHART_GAP

    # ===== 2) SubCategory =====
    ws.cell(row, 1, "2. Theo SubCategory (theo filter)").font = section_font
    row += 1
    write_header_row(row, ["SubCategory", "Chi"])
    sub_header = row
    row += 1
    sub_first = row
    for sub_name in subs:
        ws.cell(row, 1, sub_name).border = thin
        if sub_name in SUBCATEGORY_COLORS:
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=SUBCATEGORY_COLORS[sub_name])
        if sub_name == "(chưa gán)":
            sub_match = (
                f"((('{data_sheet}'!${sub_l}${data_start}:${sub_l}${data_end}=\"\")"
                f"+ISBLANK('{data_sheet}'!${sub_l}${data_start}:${sub_l}${data_end}))>0)"
            )
        else:
            sub_match = (
                f"('{data_sheet}'!${sub_l}${data_start}:${sub_l}${data_end}=A{row})"
            )
        formula = (
            f"=SUMPRODUCT({abs_amt_range()}*{month_cond}*{cat_cond}*{sub_match})"
        )
        cell = ws.cell(row, 2, formula)
        style_money(cell)
        row += 1
    sub_last = row - 1
    row += 1

    if sub_last >= sub_first:
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Chi theo SubCategory"
        bar.y_axis.title = "Số tiền"
        data = Reference(ws, min_col=2, min_row=sub_header, max_row=sub_last)
        cats_ref = Reference(ws, min_col=1, min_row=sub_first, max_row=sub_last)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.shape = 4
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        bar.width = 18
        bar.height = BAR_H
        ws.add_chart(bar, f"A{row}")
        row += rows_for_chart(BAR_H) + CHART_GAP

    # ===== 3) Month trend (lọc theo Category + SubCategory, luôn hiện mọi tháng) =====
    ws.cell(row, 1, "3. Xu hướng theo tháng (lọc Category/SubCategory)").font = section_font
    row += 1
    write_header_row(row, ["Tháng", "Chi"])
    month_header = row
    row += 1
    month_first = row
    for m in months:
        ws.cell(row, 1, m).border = thin
        formula = (
            f"=SUMPRODUCT({abs_amt_range()}*"
            f"({col_range(month_l)}=A{row})*{cat_cond}*{sub_cond})"
        )
        cell = ws.cell(row, 2, formula)
        style_money(cell)
        row += 1
    month_last = row - 1
    row += 1

    if month_last >= month_first:
        line = LineChart()
        line.style = 10
        line.title = "Chi tiêu theo tháng"
        line.y_axis.title = "Số tiền"
        line.x_axis.title = "Tháng"
        line.legend = None
        data = Reference(ws, min_col=2, min_row=month_header, max_row=month_last)
        cats_ref = Reference(ws, min_col=1, min_row=month_first, max_row=month_last)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats_ref)
        line.dataLabels = DataLabelList()
        line.dataLabels.showVal = True
        line.width = 18
        line.height = LINE_H
        ws.add_chart(line, f"A{row}")
        row += rows_for_chart(LINE_H) + CHART_GAP

    # ===== 4) Top nơi chi — snapshot tĩnh theo toàn bộ (ghi chú) =====
    # Top 10 động bằng công thức phức tạp; giữ snapshot + ghi chú filter
    spend = pd.to_numeric(df[amount_col], errors="coerce").fillna(0).abs()
    top_names = (
        pd.DataFrame({"Tên": df[name_col] if name_col else "", "Chi": spend})
        .assign(
            Tên=lambda x: x["Tên"].map(
                lambda v: "(không tên)" if _is_blank(v) else str(v).strip()
            )
        )
        .groupby("Tên")["Chi"]
        .sum()
        .sort_values(ascending=False)
        .head(10)
        .reset_index()
    )
    ws.cell(row, 1, "4. Top 10 nơi chi (toàn kỳ — không đổi theo filter)").font = section_font
    row += 1
    write_header_row(row, ["Tên định danh", "Tổng chi"])
    row += 1
    for t_row in top_names.itertuples(index=False):
        ws.cell(row, 1, t_row[0]).border = thin
        cell = ws.cell(row, 2, float(t_row[1]))
        style_money(cell)
        row += 1

    for col, width in {"A": 42, "B": 16, "C": 12, "D": 14, "E": 14, "F": 16}.items():
        ws.column_dimensions[col].width = width

    ws.sheet_view.showGridLines = False
    ws.sheet_view.topLeftCell = "A1"


if __name__ == "__main__":
    process_folder()
