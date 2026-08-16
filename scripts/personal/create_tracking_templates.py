#!/usr/bin/env python3
"""Tạo workbook Excel trống để tracking Daily / Habits / Health.

Khớp cấu trúc markdown:
  - templates/personal/daily-entry.md
  - personal/daily/README.md (ý nghĩa cột)
  - templates/personal/habit-month.md
  - personal/body/metrics.csv
  - templates/personal/nutrition-day.md

Usage:
  python scripts/personal/create_tracking_templates.py
  python scripts/personal/create_tracking_templates.py --month 2026-08
  python scripts/personal/create_tracking_templates.py --month 2026-08 --force
"""

from __future__ import annotations

import argparse
import calendar
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "data" / "personal"

# (id, tên ngắn, định nghĩa done) — mirror personal/habits/definitions.md
HABITS = [
    ("H1", "Sleep window", "Vào giường trước 23:00"),
    ("H2", "Protein morning", "Bữa sáng ≥ 30 g protein"),
    ("H3", "Deep work block", "≥1 block 90 phút không MXH"),
    ("H4", "Move", "Walk ≥20 phút hoặc gym"),
    ("H5", "No late caffeine", "Không caffeine sau 14:00"),
]

# Sheet → list[(tên cột, ý nghĩa tiếng Việt)]
SHEET_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "Daily_Meta": [
        ("Date", "Ngày (yyyy-mm-dd)"),
        ("Wake", "Giờ thức dậy, ví dụ 07:30"),
        ("Sleep_h", "Số giờ ngủ đêm trước (khớp CSV sleep_h)"),
        ("Sleep_quality_1_10", "Cảm nhận chất lượng ngủ · 1 xấu → 10 rất ngon · mục tiêu ≥6"),
        ("One_thing", "Ưu tiên / theme 1 câu của ngày — không liệt kê việc"),
    ],
    "Daily_Scores": [
        ("Date", "Ngày"),
        ("Energy", "Thể lực / tỉnh táo · 1 kiệt · 5–6 bình thường · 10 đầy"),
        ("Focus", "Giữ chú ý được bao lâu · thấp nếu cả ngày bị kéo theo thông báo"),
        ("Mood", "Cảm xúc nền · 1 xuống · 5–6 bình thường · 10 tốt"),
        ("Stress", "Áp lực việc/đời (cấp tính) · 1 êm · 8–9 rất stress"),
        ("Motivation", "Muốn làm / muốn tiến · độc lập với Energy"),
        ("Anxiety", "Lo lan tỏa, khó tắt — KHÁC Stress (Stress = áp lực việc)"),
        ("Recovery_YN", "Có nghỉ thật (walk/hobby, không chỉ ngủ)? Y = có · N = không"),
        ("Social_YN", "Có nói chuyện người thật (không chỉ chat việc)? Y/N"),
        ("Soreness_0_10", "Đau cơ / ốm · 0 = bình thường · tách khỏi stress tâm lý"),
    ],
    "Daily_DeepWork": [
        ("Date", "Ngày"),
        ("Deep_work_h", "Tổng giờ deep work THẬT (không tính mail/meeting/reactive)"),
        ("Block1_time", "Khung giờ block 1, ví dụ 09:00-10:30"),
        ("Block1_task", "Việc chính của block 1"),
        ("Block1_focus", "Độ tập trung block 1 · thang 1–10"),
        ("Block2_time", "Khung giờ block 2 (nếu có)"),
        ("Block2_task", "Việc chính của block 2"),
        ("Block2_focus", "Độ tập trung block 2 · thang 1–10"),
    ],
    "Daily_Review": [
        ("Date", "Ngày"),
        ("Wins", "Đã ship / làm xong gì hôm nay (kể cả gym)"),
        ("Friction", "Chỗ kẹt: stress, không deep, tool… → gợi ý thử tuần sau"),
        ("Note", "Ghi chú ngắn 1–3 ý"),
        ("MIT1", "Most Important Task mai — việc #1"),
        ("MIT2", "Việc #2 ngày mai (≤3 việc)"),
        ("MIT3", "Việc #3 ngày mai"),
    ],
    "Daily_Links": [
        ("Date", "Ngày"),
        (
            "Nutrition_adherence",
            "Ăn hôm nay thế nào: đầy đủ / thiếu bữa / ăn ngoài nhiều "
            "(chi tiết bữa → personal/nutrition/…)",
        ),
        (
            "Skincare_AMPM",
            "Skincare sáng-tối: đủ / thiếu bước / skip actives "
            "(chi tiết → personal/skincare/…)",
        ),
        ("CSV_filled_YN", "Đã ghi 1 dòng body/metrics.csv hôm nay? Y/N"),
        ("Nutrition_file_YN", "Đã có file nutrition ngày hôm nay? Y/N"),
        ("Habit_row_YN", "Đã tick hàng habit tháng (sheet Habits)? Y/N"),
        ("Mind_scores_YN", "Đã điền Anxiety/Recovery/Social trên Daily_Scores? Y/N"),
    ],
    "Health": [
        ("date", "Ngày — cột giống personal/body/metrics.csv"),
        ("weight_kg", "Cân nặng (kg), tùy chọn"),
        ("sleep_h", "Giờ ngủ — có số này hoặc weight → tính Metrics days"),
        ("sleep_quality_1_10", "Chất lượng ngủ 1–10"),
        ("resting_hr", "Nhịp tim nghỉ (nếu đo được), tùy chọn"),
        ("hrv_ms", "HRV mili-giây (nếu đo được), tùy chọn"),
        ("steps", "Số bước chân trong ngày"),
        (
            "training",
            "Mô tả tập (gym/run/…) — khác trống thì đếm Workouts trong Lifestyle score",
        ),
        ("waist_cm", "Vòng eo (cm), tùy chọn"),
        ("notes", "Ghi chú ngắn (giờ ngủ dậy, nước uống…)"),
    ],
    "Nutrition": [
        ("Date", "Ngày"),
        ("kcal", "Tổng kcal ước lượng trong ngày"),
        ("Protein_g", "Protein ước lượng (gram)"),
        ("Carbs_g", "Carb ước lượng (gram)"),
        ("Fat_g", "Chất béo ước lượng (gram)"),
        ("Water_ml", "Nước uống (ml)"),
        ("Hunger_1_10", "Đói trung bình trong ngày · 1 no · 10 rất đói"),
        ("Ate_late_YN", "Ăn sau 21:00? Y/N"),
        ("Alcohol_YN", "Có uống rượu/bia? Y/N"),
        ("GI_feel", "Cảm giác tiêu hóa: ổn / nặng / đầy hơi / crash sau ăn"),
        ("Repeat", "1 thứ sẽ lặp lại (ăn tốt)"),
        ("Avoid", "1 thứ sẽ tránh lần sau"),
        ("Note", "Ghi chú thêm"),
    ],
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MEANING_FILL = PatternFill("solid", fgColor="FFF2CC")
MEANING_FONT = Font(italic=True, size=9, color="595959")
SECTION_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Dòng 1 = tên cột · dòng 2 = ý nghĩa · data từ dòng 3
HEADER_ROW = 1
MEANING_ROW = 2
DATA_START = 3


def _style_header(ws, row: int = HEADER_ROW) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 30


def _style_meaning_row(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(MEANING_ROW, c)
        cell.fill = MEANING_FILL
        cell.font = MEANING_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN
    ws.row_dimensions[MEANING_ROW].height = 48


def _autosize(ws, min_w: int = 12, max_w: int = 32) -> None:
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        samples: list[str] = []
        for r in (HEADER_ROW, MEANING_ROW):
            v = ws.cell(r, idx).value
            if v:
                samples.append(str(v)[:80])
        width = max((len(s) for s in samples), default=min_w) + 2
        # ý nghĩa dài → không kéo cột quá rộng; đọc qua wrap + comment
        ws.column_dimensions[letter].width = max(min_w, min(width, max_w))


def _month_days(year: int, month: int) -> list[date]:
    n = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, n + 1)]


def _write_column_headers(ws, columns: list[tuple[str, str]]) -> int:
    """Hàng 1 tên · hàng 2 ý nghĩa + comment hover. Trả về số cột."""
    names = [c[0] for c in columns]
    meanings = [c[1] for c in columns]
    for i, (name, meaning) in enumerate(columns, start=1):
        ws.cell(HEADER_ROW, i, name)
        ws.cell(MEANING_ROW, i, meaning)
        ws.cell(HEADER_ROW, i).comment = Comment(meaning, "Docs", width=280, height=80)
    _style_header(ws, HEADER_ROW)
    _style_meaning_row(ws, len(columns))
    return len(columns)


def _fill_date_rows(ws, days: list[date], ncols: int, date_col: int = 1) -> int:
    """Ghi data từ DATA_START. Trả về chỉ số hàng cuối."""
    for i, d in enumerate(days):
        r = DATA_START + i
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
        ws.cell(r, date_col, d.isoformat())
        ws.cell(r, date_col).number_format = "yyyy-mm-dd"
    last = DATA_START + len(days) - 1
    ws.freeze_panes = "B3"  # giữ tên + ý nghĩa khi cuộn
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    return last


def _write_glossary(wb: Workbook) -> None:
    """Sheet tra cứu: mọi cột + ý nghĩa."""
    ws = wb.create_sheet("Cot_y_nghia", 0)
    ws["A1"] = "Chú thích ý nghĩa từng cột"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Mỗi sheet data cũng có hàng vàng (dòng 2) = giải thích ngắn. "
        "Di chuột vào tên cột (dòng 1) cũng hiện chú thích. "
        "Chi tiết hơn: personal/daily/README.md"
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 40

    headers = ["Sheet", "Cột", "Ý nghĩa", "Cách điền gợi ý"]
    for i, h in enumerate(headers, start=1):
        ws.cell(4, i, h)
    _style_header(ws, 4)

    # Gợi ý ngắn tách từ nghĩa (sau dấu · nếu có)
    row = 5
    for sheet, cols in SHEET_COLUMNS.items():
        for name, meaning in cols:
            tip = ""
            if "Y/N" in meaning or "Y =" in meaning:
                tip = "Chọn Y hoặc N trong dropdown"
            elif "1–10" in meaning or "1-10" in meaning or "_1_10" in name:
                tip = "Số nguyên 0–10 hoặc 1–10"
            elif name.lower() == "date" or name == "date":
                tip = "Đã điền sẵn theo tháng"
            ws.cell(row, 1, sheet)
            ws.cell(row, 2, name)
            ws.cell(row, 3, meaning)
            ws.cell(row, 4, tip)
            for c in range(1, 5):
                ws.cell(row, c).border = THIN
                ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    # Habits riêng
    ws.cell(row, 1, "Habits")
    ws.cell(row, 2, "Day / Date / H1…H5")
    ws.cell(row, 3, "Lưới habit tháng: x = làm đủ · ~ = làm một phần · - = miss")
    row += 1
    for hid, name, done in HABITS:
        ws.cell(row, 1, "Habits")
        ws.cell(row, 2, f"{hid}_{name}")
        ws.cell(row, 3, f"Done khi: {done}")
        ws.cell(row, 4, "Dropdown: x / - / ~")
        for c in range(1, 5):
            ws.cell(row, c).border = THIN
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 28
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:D{row - 1}"


def _write_legend(wb: Workbook, year: int, month: int) -> None:
    ws = wb.create_sheet("Legend", 1)
    ws["A1"] = "Personal tracking — bản đồ sheet"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    rows = [
        ("Tháng", f"{year:04d}-{month:02d}"),
        ("File", str(OUT_DIR / f"tracking-{year:04d}-{month:02d}.xlsx")),
        ("", ""),
        ("→ Đọc trước", "Sheet Cot_y_nghia = giải thích mọi cột bằng tiếng Việt"),
        ("", ""),
        ("Sheet", "Dùng để"),
        ("Daily_Meta", "Giờ thức / ngủ / việc ưu tiên 1 câu"),
        ("Daily_Scores", "Điểm thể lực–tinh thần trong ngày (thang 1–10)"),
        ("Daily_DeepWork", "Giờ làm sâu + 2 phiên"),
        ("Daily_Review", "Wins / chỗ kẹt / 3 việc mai"),
        ("Daily_Links", "Tick đã ghi nutrition/skincare/CSV chưa"),
        ("Habits", "Tick H1–H5 từng ngày trong tháng"),
        ("Health", "Số đo cơ thể — cùng cột metrics.csv"),
        ("Nutrition", "Tổng calo / macro / nước trong ngày"),
        ("", ""),
        ("Ký hiệu habit", ""),
        ("x", "Làm đủ theo định nghĩa"),
        ("~", "Làm một phần"),
        ("-", "Bỏ / miss"),
        ("(trống)", "Chưa ghi"),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)

    tip = 3 + len(rows) + 1
    ws.cell(
        tip,
        1,
        "Markdown trong personal/ vẫn là nguồn Lifestyle score "
        "(python scripts/personal_week_summary.py). Excel để ghi nhanh.",
    )
    ws.merge_cells(start_row=tip, start_column=1, end_row=tip, end_column=2)
    ws.cell(tip, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 64


def _write_daily_meta(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Daily_Meta")
    cols = SHEET_COLUMNS["Daily_Meta"]
    ncols = _write_column_headers(ws, cols)
    last = _fill_date_rows(ws, days, ncols)
    score = DataValidation(
        type="whole", operator="between", formula1="1", formula2="10", allow_blank=True
    )
    ws.add_data_validation(score)
    score.add(f"D{DATA_START}:D{last}")


def _write_daily_scores(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Daily_Scores")
    cols = SHEET_COLUMNS["Daily_Scores"]
    ncols = _write_column_headers(ws, cols)
    last = _fill_date_rows(ws, days, ncols)

    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    score = DataValidation(
        type="whole", operator="between", formula1="0", formula2="10", allow_blank=True
    )
    ws.add_data_validation(yn)
    ws.add_data_validation(score)
    score.add(f"B{DATA_START}:G{last}")
    score.add(f"J{DATA_START}:J{last}")
    yn.add(f"H{DATA_START}:I{last}")


def _write_daily_deep_work(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Daily_DeepWork")
    cols = SHEET_COLUMNS["Daily_DeepWork"]
    ncols = _write_column_headers(ws, cols)
    last = _fill_date_rows(ws, days, ncols)
    score = DataValidation(
        type="whole", operator="between", formula1="1", formula2="10", allow_blank=True
    )
    ws.add_data_validation(score)
    score.add(f"E{DATA_START}:E{last}")
    score.add(f"H{DATA_START}:H{last}")


def _write_daily_review(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Daily_Review")
    cols = SHEET_COLUMNS["Daily_Review"]
    ncols = _write_column_headers(ws, cols)
    _fill_date_rows(ws, days, ncols)
    for letter in ("B", "C", "D", "E", "F", "G"):
        ws.column_dimensions[letter].width = 28


def _write_daily_links(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Daily_Links")
    cols = SHEET_COLUMNS["Daily_Links"]
    ncols = _write_column_headers(ws, cols)
    last = _fill_date_rows(ws, days, ncols)

    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    nutr = DataValidation(
        type="list",
        formula1='"đầy đủ,thiếu bữa,ăn ngoài nhiều"',
        allow_blank=True,
    )
    skin = DataValidation(
        type="list",
        formula1='"đủ,thiếu bước,skip actives"',
        allow_blank=True,
    )
    ws.add_data_validation(yn)
    ws.add_data_validation(nutr)
    ws.add_data_validation(skin)
    nutr.add(f"B{DATA_START}:B{last}")
    skin.add(f"C{DATA_START}:C{last}")
    yn.add(f"D{DATA_START}:G{last}")


def _write_daily_sheets(wb: Workbook, days: list[date]) -> None:
    _write_daily_meta(wb, days)
    _write_daily_scores(wb, days)
    _write_daily_deep_work(wb, days)
    _write_daily_review(wb, days)
    _write_daily_links(wb, days)


def _write_habits(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Habits")
    columns = [
        ("Day", "Ngày trong tháng (1–31)"),
        ("Date", "Ngày đầy đủ"),
    ] + [
        (f"{hid}_{name}", f"Done khi: {done} · điền x / ~ / -")
        for hid, name, done in HABITS
    ]
    ncols = _write_column_headers(ws, columns)

    mark = DataValidation(type="list", formula1='"x,-,~"', allow_blank=True)
    ws.add_data_validation(mark)

    for i, d in enumerate(days):
        r = DATA_START + i
        ws.cell(r, 1, d.day)
        ws.cell(r, 2, d.isoformat())
        ws.cell(r, 2).number_format = "yyyy-mm-dd"
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
            if c >= 3:
                ws.cell(r, c).alignment = Alignment(horizontal="center")

    last = DATA_START + len(days) - 1
    mark.add(f"C{DATA_START}:{get_column_letter(ncols)}{last}")
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)

    note_row = last + 2
    ws.cell(note_row, 1, "x = làm đủ · ~ = một phần · - = miss · trống = chưa ghi")
    ws.cell(note_row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)


def _write_health(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Health")
    cols = SHEET_COLUMNS["Health"]
    ncols = _write_column_headers(ws, cols)
    last = _fill_date_rows(ws, days, ncols)
    score = DataValidation(
        type="whole", operator="between", formula1="1", formula2="10", allow_blank=True
    )
    ws.add_data_validation(score)
    score.add(f"D{DATA_START}:D{last}")

    note_row = last + 2
    ws.cell(
        note_row,
        1,
        "Có thể copy (chỉ hàng data + tên cột dòng 1) → dán vào personal/body/metrics.csv. "
        "Bỏ qua hàng vàng ý nghĩa khi copy sang CSV.",
    )
    ws.cell(note_row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)


def _write_nutrition(wb: Workbook, days: list[date]) -> None:
    ws = wb.create_sheet("Nutrition")
    cols = SHEET_COLUMNS["Nutrition"]
    ncols = _write_column_headers(ws, cols)
    last = _fill_date_rows(ws, days, ncols)

    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    score = DataValidation(
        type="whole", operator="between", formula1="1", formula2="10", allow_blank=True
    )
    gi = DataValidation(
        type="list",
        formula1='"ổn,nặng,đầy hơi,crash sau ăn"',
        allow_blank=True,
    )
    ws.add_data_validation(yn)
    ws.add_data_validation(score)
    ws.add_data_validation(gi)
    yn.add(f"H{DATA_START}:I{last}")
    score.add(f"G{DATA_START}:G{last}")
    gi.add(f"J{DATA_START}:J{last}")


def create_workbook(year: int, month: int) -> Workbook:
    days = _month_days(year, month)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    _write_glossary(wb)
    _write_legend(wb, year, month)
    _write_daily_sheets(wb, days)
    _write_habits(wb, days)
    _write_health(wb, days)
    _write_nutrition(wb, days)
    return wb


def parse_month(text: str) -> tuple[int, int]:
    dt = datetime.strptime(text.strip(), "%Y-%m")
    return dt.year, dt.month


def main() -> None:
    parser = argparse.ArgumentParser(description="Tạo Excel tracking Daily/Habits/Health")
    parser.add_argument(
        "--month",
        default=date.today().strftime("%Y-%m"),
        help="Tháng YYYY-MM (mặc định: tháng hiện tại)",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Đường dẫn file output (mặc định: data/personal/tracking-YYYY-MM.xlsx)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Ghi đè nếu file đã tồn tại",
    )
    args = parser.parse_args()

    year, month = parse_month(args.month)
    out = args.out or (OUT_DIR / f"tracking-{year:04d}-{month:02d}.xlsx")
    out = out.resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    if out.exists() and not args.force:
        raise SystemExit(f"Đã có {out} — thêm --force để ghi đè, hoặc chọn --month khác.")

    wb = create_workbook(year, month)
    wb.save(out)
    print(f"Đã tạo: {out}")
    print(
        "Sheets: Cot_y_nghia · Legend · Daily_* · Habits · Health · Nutrition"
    )
    print("Mỗi sheet data: dòng 1 = tên cột · dòng 2 (vàng) = ý nghĩa tiếng Việt")


if __name__ == "__main__":
    main()
