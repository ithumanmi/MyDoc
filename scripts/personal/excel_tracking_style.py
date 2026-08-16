"""Shared Excel formatting for personal / work tracking workbooks."""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_ROW = 1
MEANING_ROW = 2
DATA_START = 3

# --- Palette (hex không #) ---
PALETTE = {
    "navy": "1B4F72",
    "teal": "0E6655",
    "purple": "6C3483",
    "green": "196F3D",
    "orange": "B9770E",
    "slate": "34495E",
    "indigo": "4A235A",
    "crimson": "922B21",
    "sky": "1A5276",
}

SHEET_THEME: dict[str, str] = {
    # life tracking
    "Cot_y_nghia": "slate",
    "Legend": "slate",
    "Daily_Meta": "teal",
    "Daily_Scores": "teal",
    "Daily_DeepWork": "purple",
    "Daily_Review": "sky",
    "Daily_Links": "sky",
    "Habits": "purple",
    "Health": "green",
    "Nutrition": "orange",
    # work blank
    "Projects": "navy",
    "Goals": "indigo",
    "Key_Results": "indigo",
    "Timeline": "orange",
    "Prod_Daily": "teal",
    "Prod_Weekly": "sky",
    "Prod_Monthly": "green",
    "Prod_Quarterly": "purple",
    "Prod_Yearly": "crimson",
    "Perf_Weekly": "crimson",
    "Perf_Monthly": "crimson",
    "Perf_Quarterly": "crimson",
    "Perf_Yearly": "crimson",
    "Perf_Rubric": "slate",
    # farming import
    "Metrics_Log": "indigo",
}

# Giá trị → màu nền (solid)
STATUS_FILLS: dict[str, str] = {
    # project / goal / milestone
    "idea": "D5D8DC",
    "draft": "D5D8DC",
    "active": "D5F5E3",
    "paused": "FCF3CF",
    "done": "ABEBC6",
    "dropped": "F5B7B1",
    "todo": "D6EAF8",
    "doing": "F9E79F",
    "skipped": "E5E8E8",
    # metrics
    "partial": "FCF3CF",
    "open": "FADBD8",
    "shipped": "ABEBC6",
    "wip": "F9E79F",
    "blocked": "F5B7B1",
    "reverted": "E8DAEF",
    "not started": "D5D8DC",
    "on track": "D5F5E3",
    "at risk": "F5CBA7",
    "missed": "F5B7B1",
    # priority
    "p0": "F5B7B1",
    "p1": "F5CBA7",
    "p2": "FCF3CF",
    "p3": "D5D8DC",
    # effort
    "s": "D5F5E3",
    "m": "D6EAF8",
    "l": "FCF3CF",
    "xl": "F5B7B1",
    # type
    "feature": "D6EAF8",
    "fix": "FCF3CF",
    "refactor": "E8DAEF",
    "docs": "D5F5E3",
    "qa": "FDEBD0",
    # habit marks
    "x": "ABEBC6",
    "~": "F9E79F",
    "-": "F5B7B1",
}

ZEBRA_LIGHT = PatternFill("solid", fgColor="F8F9F9")
ZEBRA_ALT = PatternFill("solid", fgColor="EBF5FB")
MEANING_FILL = PatternFill("solid", fgColor="FEF9E7")
MEANING_FONT = Font(italic=True, size=9, color="7D6608", name="Calibri")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
DATA_FONT = Font(name="Calibri", size=10, color="1C2833")
TITLE_FONT = Font(bold=True, size=14, color="1B4F72", name="Calibri")
SECTION_FILL = PatternFill("solid", fgColor="D4E6F1")

# Wrap text — không tràn sang ô bên
ALIGN_WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
ALIGN_WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_WRAP_HEADER = Alignment(wrap_text=True, horizontal="center", vertical="center")

# Cột thường chứa text dài → giới hạn rộng hơn một chút nhưng vẫn wrap
_LONG_TEXT_HINTS = (
    "title",
    "note",
    "notes",
    "wins",
    "misses",
    "blocker",
    "blockers",
    "friction",
    "why",
    "deliverable",
    "next_action",
    "next_priority",
    "root_cause",
    "keep",
    "stop",
    "start",
    "energy",
    "modules",
    "quality",
    "closes",
    "blocked_by",
    "top_titles",
    "big_wins",
    "types",
    "review_note",
    "one_thing",
    "movement",
    "training",
)

THIN = Border(
    left=Side(style="thin", color="BFC9CA"),
    right=Side(style="thin", color="BFC9CA"),
    top=Side(style="thin", color="BFC9CA"),
    bottom=Side(style="thin", color="BFC9CA"),
)
MED_BOTTOM = Border(
    left=Side(style="thin", color="BFC9CA"),
    right=Side(style="thin", color="BFC9CA"),
    top=Side(style="thin", color="BFC9CA"),
    bottom=Side(style="medium", color="5D6D7E"),
)


def theme_hex(sheet_name: str) -> str:
    key = SHEET_THEME.get(sheet_name, "navy")
    return PALETTE[key]


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def status_fill(value: object) -> PatternFill | None:
    if value is None:
        return None
    key = str(value).strip().lower()
    hex_c = STATUS_FILLS.get(key)
    return fill(hex_c) if hex_c else None


def apply_tab_color(ws: Worksheet, sheet_name: str | None = None) -> None:
    name = sheet_name or ws.title
    ws.sheet_properties.tabColor = theme_hex(name)


def style_header_row(ws: Worksheet, row: int = HEADER_ROW, sheet_name: str | None = None) -> None:
    color = theme_hex(sheet_name or ws.title)
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = fill(color)
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_WRAP_HEADER
        cell.border = MED_BOTTOM
    ws.row_dimensions[row].height = 36


def style_meaning_row(ws: Worksheet, ncols: int, row: int = MEANING_ROW) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = MEANING_FILL
        cell.font = MEANING_FONT
        cell.alignment = ALIGN_WRAP_TOP
        cell.border = THIN
    ws.row_dimensions[row].height = 56


def write_column_headers(
    ws: Worksheet,
    columns: list[tuple[str, str]],
    *,
    sheet_name: str | None = None,
) -> int:
    name = sheet_name or ws.title
    for i, (col_name, meaning) in enumerate(columns, start=1):
        ws.cell(HEADER_ROW, i, col_name)
        ws.cell(MEANING_ROW, i, meaning)
        ws.cell(HEADER_ROW, i).comment = Comment(meaning, "Docs", width=300, height=90)
    style_header_row(ws, HEADER_ROW, name)
    style_meaning_row(ws, len(columns))
    apply_tab_color(ws, name)
    return len(columns)


def paint_data_row(
    ws: Worksheet,
    row: int,
    ncols: int,
    *,
    zebra_index: int = 0,
) -> None:
    base = ZEBRA_ALT if zebra_index % 2 else ZEBRA_LIGHT
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.border = THIN
        cell.font = DATA_FONT
        if cell.fill is None or cell.fill.fgColor is None or cell.fill.fgColor.rgb in (
            "00000000",
            None,
        ):
            cell.fill = base
        # only set zebra if no status fill yet — check patternType
        if cell.fill.patternType is None:
            cell.fill = base


def paint_row_zebra(ws: Worksheet, row: int, ncols: int, zebra_index: int) -> None:
    base = ZEBRA_ALT if zebra_index % 2 else ZEBRA_LIGHT
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = base
        cell.border = THIN
        cell.font = DATA_FONT
        cell.alignment = ALIGN_WRAP_TOP


def paint_status_cell(cell, value: object | None = None) -> None:
    v = value if value is not None else cell.value
    f = status_fill(v)
    if f:
        cell.fill = f
        cell.alignment = ALIGN_WRAP_CENTER
        cell.font = Font(name="Calibri", size=10, bold=True, color="1C2833")


def finish_data_sheet(
    ws: Worksheet,
    ncols: int,
    last_row: int,
    *,
    freeze: str = "B3",
) -> None:
    ws.freeze_panes = freeze
    if last_row >= HEADER_ROW:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last_row}"
    ws.sheet_view.showGridLines = False
    autosize(ws)
    apply_wrap_no_overflow(ws, DATA_START, last_row, ncols)


def _is_long_text_col(header: str | None) -> bool:
    if not header:
        return False
    h = str(header).strip().lower()
    return any(k in h for k in _LONG_TEXT_HINTS)


def apply_wrap_no_overflow(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    ncols: int,
) -> None:
    """Bật wrap text + chiều cao hàng — chữ không tràn ô cạnh."""
    if end_row < start_row or ncols < 1:
        return
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, ncols + 1)]
    for r in range(start_row, end_row + 1):
        max_lines = 1
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            if status_fill(cell.value):
                cell.alignment = ALIGN_WRAP_CENTER
            else:
                cell.alignment = ALIGN_WRAP_TOP
            val = cell.value
            if val is None:
                continue
            text = str(val)
            col_w = ws.column_dimensions[get_column_letter(c)].width or 12
            chars_per_line = max(int(col_w), 8)
            soft = (len(text) + chars_per_line - 1) // chars_per_line
            hard = text.count("\n") + 1
            lines = max(soft, hard)
            if _is_long_text_col(headers[c - 1] if c - 1 < len(headers) else None):
                lines = max(lines, min((len(text) // 24) + 1, 6))
            max_lines = max(max_lines, min(lines, 8))
        ws.row_dimensions[r].height = max(18, min(15 * max_lines, 90))


def autosize(ws: Worksheet, min_w: int = 10, max_w: int = 28) -> None:
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, (ws.max_column or 0) + 1)]
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        samples: list[str] = []
        for r in (HEADER_ROW, MEANING_ROW):
            v = ws.cell(r, idx).value
            if v:
                samples.append(str(v)[:40])
        for r in range(DATA_START, min(DATA_START + 8, (ws.max_row or DATA_START) + 1)):
            v = ws.cell(r, idx).value
            if v is not None:
                samples.append(str(v)[:28])
        width = max((len(s) for s in samples), default=min_w) + 2
        hdr = headers[idx - 1] if idx - 1 < len(headers) else None
        cap = 22 if _is_long_text_col(hdr) else max_w
        h = str(hdr or "").lower()
        if h in {"date", "week", "month", "year", "day", "quarter"} or h.endswith(
            "_1_10"
        ) or h.endswith("_pct") or h in {"items", "done", "shipped", "reopen", "reopen_sum"}:
            cap = min(cap, 14)
        ws.column_dimensions[letter].width = max(min_w, min(width, cap))


def add_status_conditional_formats(
    ws: Worksheet,
    col_letter: str,
    start_row: int,
    end_row: int,
    values: list[str] | None = None,
) -> None:
    """Khi user chọn dropdown, ô đổi màu theo giá trị."""
    if end_row < start_row:
        return
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    keys = values or list(STATUS_FILLS.keys())
    for key in keys:
        hex_c = STATUS_FILLS.get(key)
        if not hex_c:
            continue
        formula = f'UPPER({col_letter}{start_row})="{key.upper()}"'
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[formula],
                fill=fill(hex_c),
                font=Font(bold=True, name="Calibri", size=10),
            ),
        )


def add_score_color_scale_hint(
    ws: Worksheet,
    col_letter: str,
    start_row: int,
    end_row: int,
) -> None:
    """Tô điểm 1–10: thấp đỏ · cao xanh (formula bands)."""
    if end_row < start_row:
        return
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    bands = [
        (1, 3, "F5B7B1"),
        (4, 6, "FCF3CF"),
        (7, 8, "D5F5E3"),
        (9, 10, "82E0AA"),
    ]
    for lo, hi, hex_c in bands:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=[str(lo), str(hi)],
                fill=fill(hex_c),
            ),
        )


def style_title_block(ws: Worksheet, cell: str = "A1", text: str = "") -> None:
    c = ws[cell]
    if text:
        c.value = text
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center")


def style_legend_sheet(ws: Worksheet) -> None:
    apply_tab_color(ws, "Legend")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 68
    swatch_start = (ws.max_row or 20) + 3
    ws.cell(swatch_start, 1, "Bảng màu trạng thái").font = Font(bold=True, size=11)
    samples = [
        ("done / shipped / x", "done"),
        ("partial / doing / ~", "partial"),
        ("open / blocked / -", "open"),
        ("Active / on track", "active"),
        ("P0", "p0"),
        ("Feature", "feature"),
        ("Fix", "fix"),
        ("XL effort", "xl"),
    ]
    for i, (label, key) in enumerate(samples):
        r = swatch_start + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 1).fill = fill(STATUS_FILLS[key])
        ws.cell(r, 1).border = THIN


# Cột thường dùng dropdown / status → conditional format
_STATUS_COL_HINTS = {
    "status",
    "outcome",
    "priority",
    "effort",
    "type",
    "horizon",
}
_SCORE_COL_HINTS = {
    "score_1_10",
    "throughput_1_10",
    "quality_1_10",
    "focus_1_10",
    "stability_1_10",
    "overall_1_10",
    "energy",
    "focus",
    "mood",
    "stress",
    "motivation",
    "anxiety",
    "soreness_0_10",
    "sleep_quality_1_10",
    "hunger_1_10",
    "block1_focus",
    "block2_focus",
    "energy_work_1_10",
    "complexity",
}


def _header_map(ws: Worksheet) -> dict[str, int]:
    """Tên cột (lower) → index 1-based từ hàng HEADER_ROW."""
    out: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(HEADER_ROW, c).value
        if v:
            out[str(v).strip().lower()] = c
    return out


def apply_zebra_band(ws: Worksheet, start_row: int, end_row: int, ncols: int) -> None:
    for i, r in enumerate(range(start_row, end_row + 1)):
        base = ZEBRA_ALT if i % 2 else ZEBRA_LIGHT
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            # giữ fill status đã tô (màu không phải zebra)
            fg = None
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                fg = str(cell.fill.fgColor.rgb).upper().lstrip("FF")
            if fg and fg not in {"F8F9F9", "EBF5FB", "00000000", "000000"}:
                # đã có màu khác → chỉ border/font
                cell.border = THIN
                if not cell.font or cell.font.name is None:
                    cell.font = DATA_FONT
                cell.alignment = ALIGN_WRAP_TOP
                continue
            cell.fill = base
            cell.border = THIN
            cell.font = DATA_FONT
            cell.alignment = ALIGN_WRAP_TOP


def paint_cells_by_value(
    ws: Worksheet,
    col_idx: int,
    start_row: int,
    end_row: int,
) -> None:
    for r in range(start_row, end_row + 1):
        cell = ws.cell(r, col_idx)
        paint_status_cell(cell)


def polish_workbook(wb) -> None:
    """Áp theme tab, zebra, conditional format, màu status đã có giá trị."""
    for ws in wb.worksheets:
        name = ws.title
        apply_tab_color(ws, name)
        ws.sheet_view.showGridLines = False

        if name in {"Legend", "Cot_y_nghia"}:
            if name == "Legend":
                style_legend_sheet(ws)
            else:
                apply_tab_color(ws, "Cot_y_nghia")
            # restyle title
            if ws["A1"].value:
                style_title_block(ws, "A1")
            continue

        ncols = ws.max_column or 0
        max_r = ws.max_row or 0
        if ncols < 1 or max_r < HEADER_ROW:
            continue

        # Header theme color
        if ws.cell(HEADER_ROW, 1).value:
            style_header_row(ws, HEADER_ROW, name)
        if max_r >= MEANING_ROW and ws.cell(MEANING_ROW, 1).value:
            style_meaning_row(ws, ncols)

        last = max_r
        if last >= DATA_START:
            apply_zebra_band(ws, DATA_START, last, ncols)

        headers = _header_map(ws)
        for hname, col in headers.items():
            letter = get_column_letter(col)
            # habit marks H1_*
            if hname.startswith("h1_") or hname.startswith("h2_") or hname.startswith(
                "h3_"
            ) or hname.startswith("h4_") or hname.startswith("h5_"):
                add_status_conditional_formats(
                    ws, letter, DATA_START, last, values=["x", "-", "~"]
                )
                paint_cells_by_value(ws, col, DATA_START, last)
                continue

            base = hname.split("_")[0] if "_" in hname else hname
            if hname in _STATUS_COL_HINTS or base in _STATUS_COL_HINTS:
                add_status_conditional_formats(ws, letter, DATA_START, last)
                paint_cells_by_value(ws, col, DATA_START, last)
            if hname in _SCORE_COL_HINTS or hname.endswith("_1_10") or hname.endswith(
                "1_10"
            ):
                add_score_color_scale_hint(ws, letter, DATA_START, last)

        # freeze + filter if data sheet
        if last >= DATA_START and ws.freeze_panes is None:
            ws.freeze_panes = "B3"
        if last >= HEADER_ROW:
            try:
                ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
            except Exception:
                pass
        autosize(ws)
        if last >= DATA_START:
            apply_wrap_no_overflow(ws, DATA_START, last, ncols)
            # meaning row cũng wrap chặt
            apply_wrap_no_overflow(ws, MEANING_ROW, MEANING_ROW, ncols)
            ws.row_dimensions[MEANING_ROW].height = max(
                ws.row_dimensions[MEANING_ROW].height or 56, 56
            )
