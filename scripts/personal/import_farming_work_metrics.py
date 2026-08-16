#!/usr/bin/env python3
"""Import Farming Docs (WORK_METRICS + STATUS + epic) → work-tracking Excel.

Nguồn mặc định:
  C:/Projects/Game/Games/Farming/Docs/WORK_METRICS.md
  STATUS.md · BACKLOG.md · production/epics/.../EPIC.md

Output:
  data/personal/work-tracking-farming-YYYY.xlsx

Sheets:
  Metrics_Log  — dump 1:1 từ WORK_METRICS (nguồn phân tích)
  Prod_*       — roll-up ngày/tuần/tháng/quý/năm
  Perf_*       — đánh giá performance (điểm + Keep/Stop/Start)
  Perf_Rubric  — rubric thang điểm
  Projects     — hệ thống từ STATUS.md
  Timeline     — story epic + Open P0/P1 từ BACKLOG
  Goals        — 1 goal dự án Farming + focus gần đây

Usage:
  python scripts/personal/import_farming_work_metrics.py
  python scripts/personal/import_farming_work_metrics.py --year 2026 --force
  python scripts/personal/import_farming_work_metrics.py --farming-docs "C:/Projects/Game/Games/Farming/Docs"
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_tracking_style import polish_workbook  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "data" / "personal"
DEFAULT_FARMING_DOCS = Path(r"C:\Projects\Game\Games\Farming\Docs")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MEANING_FILL = PatternFill("solid", fgColor="FFF2CC")
MEANING_FONT = Font(italic=True, size=9, color="595959")
SECTION_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

HEADER_ROW = 1
MEANING_ROW = 2
DATA_START = 3

# Ước giờ từ thang effort WORK_METRICS (không phải time tracking thật)
EFFORT_HOURS = {"S": 1.5, "M": 4.0, "L": 8.0, "XL": 16.0}

METRICS_COLS = [
    ("date", "Ngày ghi (khớp DAILY)"),
    ("week", "Tuần ISO"),
    ("month", "Tháng YYYY-MM"),
    ("title", "Tên cụm việc (rút gọn DAILY)"),
    ("type", "Feature / Fix / Refactor / Docs / QA"),
    ("category", "Nhóm việc: Camera, Feel, Editor UX…"),
    ("modules", "SYSTEMS liên quan (; tách)"),
    ("effort", "S/M/L/XL — quy mô ước lượng"),
    ("complexity", "1–5 độ khó"),
    ("status", "done / partial / open"),
    ("outcome", "shipped / wip / blocked / reverted"),
    ("quality", "tests_ok; needs_playtest; docs_synced…"),
    ("reopen", "Số lần đụng lại chủ đề"),
    ("blocked_by", "Token treo (không phải backlog thường)"),
    ("closes", "Đóng nợ cụm trước"),
    ("notes", "1 ý ngắn cho báo cáo"),
    ("story_id", "ID story/epic (PS-002) hoặc —"),
    ("priority", "P0–P3 hoặc —"),
    ("hours_actual", "Giờ thật nếu có; — nếu không"),
    ("verify", "editmode_ok / playmode_pending / playmode_ok / unverified"),
    ("deep_work", "1 = deep work · 0 = không"),
    ("est_hours", "Giờ dùng báo cáo: hours_actual nếu có, không thì map effort"),
]

PROJECT_ID = "P-farming"
GOAL_ID = "G-farming-2026"


def _style_header(ws) -> None:
    for cell in ws[HEADER_ROW]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[HEADER_ROW].height = 30


def _style_meaning(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(MEANING_ROW, c)
        cell.fill = MEANING_FILL
        cell.font = MEANING_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN
    ws.row_dimensions[MEANING_ROW].height = 48


def _write_headers(ws, columns: list[tuple[str, str]]) -> int:
    for i, (name, meaning) in enumerate(columns, start=1):
        ws.cell(HEADER_ROW, i, name)
        ws.cell(MEANING_ROW, i, meaning)
        ws.cell(HEADER_ROW, i).comment = Comment(meaning, "Docs", width=280, height=80)
    _style_header(ws)
    _style_meaning(ws, len(columns))
    return len(columns)


def _autosize(ws, max_w: int = 22) -> None:
    """Giới hạn rộng cột — text dài sẽ wrap thay vì kéo ngang / tràn."""
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        samples = []
        for r in (HEADER_ROW, MEANING_ROW, DATA_START):
            v = ws.cell(r, idx).value
            if v:
                samples.append(str(v)[:24])
        w = max((len(s) for s in samples), default=10) + 2
        ws.column_dimensions[letter].width = min(max(w, 9), max_w)


def _split_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [c.strip() for c in line.split("|")]


def parse_metrics_table(path: Path) -> list[dict[str, str]]:
    text = path.read_text(encoding="utf-8")
    # Tìm bảng Data: header có date | week | month | title
    lines = text.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if re.match(r"^\|\s*date\s*\|", line, re.I):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"Không tìm thấy bảng metrics trong {path}")

    headers = [h.lower() for h in _split_row(lines[header_idx])]
    rows: list[dict[str, str]] = []
    for line in lines[header_idx + 1 :]:
        if not line.strip().startswith("|"):
            if rows:
                break
            continue
        # separator | --- |
        if re.match(r"^\|\s*:?-{3,}", line):
            continue
        cells = _split_row(line)
        if len(cells) < 4:
            continue
        # pad/truncate
        while len(cells) < len(headers):
            cells.append("")
        row = {headers[i]: cells[i] for i in range(len(headers))}
        if not row.get("date") or row["date"].startswith("---"):
            continue
        rows.append(row)
    return rows


def effort_hours(effort: str) -> float | None:
    key = (effort or "").strip().upper()
    if key in EFFORT_HOURS:
        return EFFORT_HOURS[key]
    return None


def row_hours(row: dict[str, str]) -> float:
    """Ưu tiên hours_actual; không thì map effort → giờ ước."""
    raw = (row.get("hours_actual") or "").strip().replace(",", ".")
    if raw and raw not in {"—", "-", "–"}:
        try:
            return float(raw)
        except ValueError:
            pass
    return effort_hours(row.get("effort", "")) or 0.0


def blankish(val: str | None) -> str:
    v = (val or "").strip()
    return "" if v in {"", "—", "-", "–"} else v


def parse_status_projects(status_path: Path) -> list[dict[str, str]]:
    if not status_path.exists():
        return []
    text = status_path.read_text(encoding="utf-8")
    # Bảng ## Systems
    m = re.search(r"## Systems\s*\n\n(\|.+\n\|.+\n(?:\|.+\n)+)", text)
    if not m:
        return []
    block = m.group(1).strip().splitlines()
    headers = [h.lower() for h in _split_row(block[0])]
    projects = []
    for line in block[2:]:
        if not line.startswith("|"):
            break
        cells = _split_row(line)
        if len(cells) < 3:
            continue
        data = {headers[i]: cells[i] if i < len(cells) else "" for i in range(len(headers))}
        system = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", data.get("system", "")).strip()
        status = data.get("status", "").strip()
        debt = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", data.get("nợ chính (≤3)", data.get("nợ chính", ""))).strip()
        map_status = {
            "Stable": "Active",
            "WIP": "Active",
            "Thin": "Paused",
            "Partial": "Paused",
            "Stub": "Idea",
            "Absent": "Dropped",
        }
        # take first word of status tag
        st_key = status.split()[0] if status else "WIP"
        projects.append(
            {
                "Project_ID": f"P-{re.sub(r'[^a-zA-Z0-9]+', '-', system).strip('-').lower()[:40]}",
                "Name": system,
                "Status": map_status.get(st_key, "Active"),
                "Priority": "P1" if st_key == "WIP" else ("P2" if st_key == "Thin" else "P3"),
                "Start": "",
                "End": "",
                "Goal_ID": GOAL_ID,
                "Progress_pct": "",
                "Next_action": debt[:200] if debt else "",
                "Note": f"STATUS tag: {status}",
            }
        )
    # parent project first
    projects.insert(
        0,
        {
            "Project_ID": PROJECT_ID,
            "Name": "Farming (Unity game)",
            "Status": "Active",
            "Priority": "P0",
            "Start": "2026-07-22",
            "End": "",
            "Goal_ID": GOAL_ID,
            "Progress_pct": "",
            "Next_action": "PlayMode smoke Sequence 002–007; killing-blow 001",
            "Note": f"Docs: {status_path.parent}",
        },
    )
    return projects


def parse_backlog_timeline(backlog_path: Path) -> list[dict[str, str]]:
    if not backlog_path.exists():
        return []
    text = backlog_path.read_text(encoding="utf-8")
    m = re.search(r"## Open \(ưu tiên theo focus hiện tại\)\s*\n\n(\|.+\n\|.+\n(?:\|.+\n)+)", text)
    if not m:
        return []
    block = m.group(1).strip().splitlines()
    headers = [h.lower() for h in _split_row(block[0])]
    items = []
    for i, line in enumerate(block[2:], start=1):
        if not line.startswith("|"):
            break
        cells = _split_row(line)
        data = {headers[j]: cells[j] if j < len(cells) else "" for j in range(len(headers))}
        pri = data.get("priority", "P2").strip()
        system = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", data.get("system", "")).strip()
        open_txt = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", data.get("open (rút gọn)", "")).strip()
        items.append(
            {
                "Milestone_ID": f"BL-{i:02d}",
                "Project_ID": PROJECT_ID,
                "Goal_ID": GOAL_ID,
                "Title": f"[{pri}] {system}: {open_txt}"[:120],
                "Due_date": "",
                "Status": "Todo",
                "Deliverable": open_txt[:160],
                "Note": "Từ BACKLOG.md Open",
            }
        )
    return items


def parse_epic_timeline(epic_path: Path) -> list[dict[str, str]]:
    if not epic_path.exists():
        return []
    text = epic_path.read_text(encoding="utf-8")
    items = []
    # | ID | File | Wave | Priority | Status | ...
    for line in text.splitlines():
        if not re.match(r"^\|\s*\d{3}\s*\|", line):
            continue
        cells = _split_row(line)
        if len(cells) < 5:
            continue
        sid, _file, wave, priority, status = cells[0], cells[1], cells[2], cells[3], cells[4]
        title = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", _file).strip()
        st = status.lower()
        map_st = "Done" if "done" in st else ("Doing" if "progress" in st or "in-progress" in st else "Todo")
        items.append(
            {
                "Milestone_ID": f"PS-{sid}",
                "Project_ID": "P-presentation-sequence",
                "Goal_ID": GOAL_ID,
                "Title": f"Story {sid}: {title}"[:120],
                "Due_date": "",
                "Status": map_st,
                "Deliverable": f"Wave {wave} · {priority}",
                "Note": "Từ presentation-sequence EPIC.md",
            }
        )
    return items


def parse_daily_metrics_reviews(daily_path: Path) -> dict[str, dict[str, str]]:
    """Parse ### Metrics review trong ## Tổng kết tuần/tháng/quý.

    Key: YYYY-Wnn | YYYY-MM | YYYY-Qn
    """
    if not daily_path.exists():
        return {}
    text = daily_path.read_text(encoding="utf-8")
    out: dict[str, dict[str, str]] = {}
    # Split by ## headings
    parts = re.split(r"(?m)^(## .+)$", text)
    # parts: preamble, heading, body, heading, body...
    i = 1
    while i + 1 < len(parts):
        heading = parts[i].strip()
        body = parts[i + 1]
        i += 2
        period = None
        m = re.search(r"Tổng kết tuần\s+(20\d{2}-W\d{2})", heading, re.I)
        if m:
            period = m.group(1)
        else:
            m = re.search(r"Tổng kết tháng\s+(20\d{2}-\d{2})", heading, re.I)
            if m:
                period = m.group(1)
            else:
                m = re.search(r"Tổng kết quý\s+(20\d{2}-Q[1-4])", heading, re.I)
                if m:
                    period = m.group(1)
        if not period:
            continue
        block = re.search(r"###\s*Metrics review\s*\n((?:[-*].+\n?)+)", body, re.I)
        if not block:
            continue
        fields: dict[str, str] = {}
        for line in block.group(1).splitlines():
            line = line.lstrip("-* ").strip()
            if ":" not in line:
                continue
            key, _, val = line.partition(":")
            key = key.strip().lower()
            val = val.strip()
            if key.startswith("energy"):
                # energy: 7 · overall: 8
                em = re.search(r"(\d+|—|-)", val)
                fields["Energy"] = em.group(1) if em else val
                om = re.search(r"overall:\s*(\d+|—|-)", val, re.I)
                if om:
                    fields["Overall_1_10"] = om.group(1)
            elif key == "overall":
                fields["Overall_1_10"] = re.sub(r"[^\d—-]", "", val) or val
            elif key == "wins":
                fields["Wins"] = val
            elif key == "misses":
                fields["Misses"] = val
            elif key == "keep":
                fields["Keep"] = val
            elif key == "stop":
                fields["Stop"] = val
            elif key == "start":
                fields["Start"] = val
            elif key in {"next", "next_priority"}:
                fields["Next_priority"] = val
        if fields:
            out[period] = fields
    return out


def aggregate_productivity(rows: list[dict[str, str]], year: int) -> dict[str, list[dict]]:
    by_day: dict[str, list] = defaultdict(list)
    by_week: dict[str, list] = defaultdict(list)
    by_month: dict[str, list] = defaultdict(list)

    for r in rows:
        d = r.get("date", "")
        if not d.startswith(str(year)):
            # vẫn nhận năm khác nếu filter rộng — skip ngoài year
            try:
                if datetime.strptime(d, "%Y-%m-%d").year != year:
                    continue
            except ValueError:
                continue
        by_day[d].append(r)
        by_week[r.get("week", "")].append(r)
        by_month[r.get("month", "")].append(r)

    def pack(group: list[dict], label_key: str, label: str) -> dict:
        hours = sum(row_hours(x) for x in group)
        done = sum(1 for x in group if x.get("status") == "done")
        shipped = sum(1 for x in group if x.get("outcome") == "shipped")
        deep = sum(1 for x in group if blankish(x.get("deep_work")) == "1")
        p01 = sum(1 for x in group if blankish(x.get("priority")).upper() in {"P0", "P1"})
        types = defaultdict(int)
        for x in group:
            types[x.get("type", "?")] += 1
        blockers = []
        for x in group:
            b = blankish(x.get("blocked_by"))
            if b:
                blockers.append(b)
        titles = [x.get("title", "") for x in group[:3]]
        return {
            label_key: label,
            "items": len(group),
            "done": done,
            "shipped": shipped,
            "est_hours": round(hours, 1),
            "deep_work_n": deep,
            "p01_n": p01,
            "types": "; ".join(f"{k}:{v}" for k, v in sorted(types.items())),
            "top_titles": " · ".join(titles),
            "blockers": "; ".join(dict.fromkeys(blockers))[:200],
        }

    daily = [pack(v, "Date", k) for k, v in sorted(by_day.items())]
    weekly = []
    for week, group in sorted(by_week.items()):
        if not week:
            continue
        p = pack(group, "Week", week)
        # week start Monday
        try:
            # parse YYYY-Www
            y, w = week.split("-W")
            monday = date.fromisocalendar(int(y), int(w), 1)
            p["Week_start"] = monday.isoformat()
        except ValueError:
            p["Week_start"] = ""
        weekly.append(p)
    monthly = [pack(v, "Month", k) for k, v in sorted(by_month.items()) if k]

    # quarter
    by_q: dict[str, list] = defaultdict(list)
    for r in rows:
        d = r.get("date", "")
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            continue
        if dt.year != year:
            continue
        q = (dt.month - 1) // 3 + 1
        by_q[f"{year}-Q{q}"].append(r)
    quarterly = [pack(v, "Quarter", k) for k, v in sorted(by_q.items())]

    year_rows = [r for r in rows if r.get("date", "").startswith(str(year))]
    yearly = [pack(year_rows, "Year", str(year))] if year_rows else []

    return {
        "daily": daily,
        "weekly": weekly,
        "monthly": monthly,
        "quarterly": quarterly,
        "yearly": yearly,
    }


def _write_sheet_rows(ws, columns: list[tuple[str, str]], records: list[dict], key_order: list[str]) -> None:
    ncols = _write_headers(ws, columns)
    for i, rec in enumerate(records):
        r = DATA_START + i
        for c, key in enumerate(key_order, start=1):
            val = rec.get(key, "")
            cell = ws.cell(r, c, val if val != "" else None)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if key in {"date", "Date", "Start", "End", "Due_date", "Week_start"} and val:
                cell.number_format = "yyyy-mm-dd"
    last = DATA_START + max(len(records), 1) - 1
    if not records:
        for c in range(1, ncols + 1):
            ws.cell(DATA_START, c).border = THIN
            ws.cell(DATA_START, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)


def build_workbook(
    rows: list[dict[str, str]],
    year: int,
    projects: list[dict],
    timeline: list[dict],
    farming_docs: Path,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # Legend
    leg = wb.create_sheet("Legend", 0)
    leg["A1"] = "Farming → Work tracking (import từ Docs game)"
    leg["A1"].font = Font(bold=True, size=14)
    info = [
        ("Nguồn metrics", str(farming_docs / "WORK_METRICS.md")),
        ("Nguồn STATUS", str(farming_docs / "STATUS.md")),
        ("Nguồn BACKLOG", str(farming_docs / "BACKLOG.md")),
        ("DAILY (narrative)", str(farming_docs / "DAILY.md")),
        ("Số hàng Metrics_Log", str(len(rows))),
        ("Năm filter Prod_*", str(year)),
        ("", ""),
        ("Cách đọc", ""),
        ("Metrics_Log", "Dump 1:1 WORK_METRICS — dùng để pivot / lọc"),
        ("Prod_*", "Roll-up: giờ = hours_actual nếu có, không thì map effort S/M/L/XL"),
        ("Projects", "Hệ thống từ STATUS + parent P-farming"),
        ("Timeline", "Epic stories + Open BACKLOG"),
        ("Goals", "Mục tiêu dự án Farming năm"),
        ("Perf_*", "Snapshot metrics + Metrics review từ DAILY (nếu có); còn trống thì điền tay"),
        ("Perf_Rubric", "Rubric giải thích thang điểm"),
        ("", ""),
        ("Lưu ý", "Hàng cũ thiếu story_id/priority/hours_actual/verify/deep_work → coi như —"),
        ("Cập nhật", "Sửa WORK_METRICS/DAILY trong game Docs rồi chạy lại script --force"),
    ]
    for i, (a, b) in enumerate(info, start=3):
        leg.cell(i, 1, a)
        leg.cell(i, 2, b)
    leg.column_dimensions["A"].width = 22
    leg.column_dimensions["B"].width = 70

    reviews = parse_daily_metrics_reviews(farming_docs / "DAILY.md")

    # Metrics_Log
    metrics_records = []
    for r in rows:
        item = {k: r.get(k, "") for k, _ in METRICS_COLS if k != "est_hours"}
        h = row_hours(r)
        # nếu không có effort lẫn hours_actual → để trống thay vì 0
        has_actual = blankish(r.get("hours_actual")) != ""
        has_effort = effort_hours(r.get("effort", "")) is not None
        item["est_hours"] = round(h, 1) if (has_actual or has_effort) else ""
        metrics_records.append(item)
    # newest first already in file — keep order
    ws = wb.create_sheet("Metrics_Log")
    _write_sheet_rows(ws, METRICS_COLS, metrics_records, [c[0] for c in METRICS_COLS])

    agg = aggregate_productivity(rows, year)

    # Prod_Daily
    daily_cols = [
        ("Date", "Ngày có ít nhất 1 cụm trong WORK_METRICS"),
        ("Items", "Số cụm việc ghi trong ngày"),
        ("Done", "Số status=done"),
        ("Shipped", "Số outcome=shipped"),
        ("Est_hours", "Tổng giờ: hours_actual ưu tiên, không thì map effort"),
        ("Types", "Đếm theo Feature/Fix/…"),
        ("Top_titles", "≤3 title đầu"),
        ("Blockers", "blocked_by gộp"),
        ("Project_ID", "Mặc định P-farming"),
    ]
    daily_recs = [
        {
            "Date": x["Date"],
            "Items": x["items"],
            "Done": x["done"],
            "Shipped": x["shipped"],
            "Est_hours": x["est_hours"],
            "Types": x["types"],
            "Top_titles": x["top_titles"],
            "Blockers": x["blockers"],
            "Project_ID": PROJECT_ID,
        }
        for x in agg["daily"]
    ]
    _write_sheet_rows(wb.create_sheet("Prod_Daily"), daily_cols, daily_recs, [c[0] for c in daily_cols])

    weekly_cols = [
        ("Week", "Tuần ISO"),
        ("Week_start", "Thứ Hai"),
        ("Items", "Số cụm"),
        ("Done", "done"),
        ("Shipped", "shipped"),
        ("Est_hours", "Giờ ước"),
        ("Types", "Phân bố type"),
        ("Big_wins", "Top titles"),
        ("Blockers", "blocked_by"),
        ("Score_1_10", "Để trống — tự chấm khi review"),
    ]
    weekly_recs = [
        {
            "Week": x["Week"],
            "Week_start": x.get("Week_start", ""),
            "Items": x["items"],
            "Done": x["done"],
            "Shipped": x["shipped"],
            "Est_hours": x["est_hours"],
            "Types": x["types"],
            "Big_wins": x["top_titles"],
            "Blockers": x["blockers"],
            "Score_1_10": "",
        }
        for x in agg["weekly"]
    ]
    _write_sheet_rows(wb.create_sheet("Prod_Weekly"), weekly_cols, weekly_recs, [c[0] for c in weekly_cols])

    monthly_cols = [
        ("Month", "YYYY-MM"),
        ("Items", "Số cụm"),
        ("Done", "done"),
        ("Shipped", "shipped"),
        ("Est_hours", "Giờ ước"),
        ("Types", "Phân bố"),
        ("Wins", "Top titles"),
        ("Blockers", "blocked_by"),
        ("Score_1_10", "Tự chấm"),
    ]
    monthly_recs = [
        {
            "Month": x["Month"],
            "Items": x["items"],
            "Done": x["done"],
            "Shipped": x["shipped"],
            "Est_hours": x["est_hours"],
            "Types": x["types"],
            "Wins": x["top_titles"],
            "Blockers": x["blockers"],
            "Score_1_10": "",
        }
        for x in agg["monthly"]
    ]
    _write_sheet_rows(wb.create_sheet("Prod_Monthly"), monthly_cols, monthly_recs, [c[0] for c in monthly_cols])

    q_cols = [
        ("Quarter", "YYYY-Qn"),
        ("Items", "Số cụm"),
        ("Done", "done"),
        ("Shipped", "shipped"),
        ("Est_hours", "Giờ ước"),
        ("KR_hit_pct", "Ước = shipped/items*100 nếu items>0"),
        ("Wins", "Top titles"),
        ("Misses", "Blockers gộp"),
        ("Score_1_10", "Tự chấm"),
    ]
    q_recs = []
    for x in agg["quarterly"]:
        pct = round(100 * x["shipped"] / x["items"]) if x["items"] else ""
        q_recs.append(
            {
                "Quarter": x["Quarter"],
                "Items": x["items"],
                "Done": x["done"],
                "Shipped": x["shipped"],
                "Est_hours": x["est_hours"],
                "KR_hit_pct": pct,
                "Wins": x["top_titles"],
                "Misses": x["blockers"],
                "Score_1_10": "",
            }
        )
    _write_sheet_rows(wb.create_sheet("Prod_Quarterly"), q_cols, q_recs, [c[0] for c in q_cols])

    y_cols = [
        ("Year", "Năm"),
        ("Items", "Tổng cụm"),
        ("Done", "done"),
        ("Shipped", "shipped"),
        ("Est_hours", "Giờ ước"),
        ("Top_wins", "Title mẫu"),
        ("Blockers", "blocked_by"),
        ("Score_1_10", "Tự chấm"),
    ]
    y_recs = [
        {
            "Year": x["Year"],
            "Items": x["items"],
            "Done": x["done"],
            "Shipped": x["shipped"],
            "Est_hours": x["est_hours"],
            "Top_wins": x["top_titles"],
            "Blockers": x["blockers"],
            "Score_1_10": "",
        }
        for x in agg["yearly"]
    ]
    _write_sheet_rows(wb.create_sheet("Prod_Yearly"), y_cols, y_recs, [c[0] for c in y_cols])

    # --- Performance review (điền tay; snapshot số từ metrics) ---
    from perf_review_schema import PERF_RUBRIC_ROWS, perf_columns

    def _reopen_sum(group_rows: list[dict]) -> int:
        total = 0
        for r in group_rows:
            raw = (r.get("reopen") or "0").strip()
            if raw in {"", "—", "-"}:
                continue
            try:
                total += int(float(raw))
            except ValueError:
                pass
        return total

    def _perf_rec(period_key: str, period: str, snap: dict, group_rows: list[dict] | None = None) -> dict:
        items = snap["items"]
        shipped = snap["shipped"]
        ship_pct = round(100 * shipped / items) if items else ""
        rev = reviews.get(period, {})
        wins = blankish(rev.get("Wins")) or snap["top_titles"]
        return {
            period_key: period,
            "Items": items,
            "Done": snap["done"],
            "Shipped": shipped,
            "Ship_pct": ship_pct,
            "Est_hours": snap["est_hours"],
            "Reopen_sum": _reopen_sum(group_rows or []),
            "Blockers": snap["blockers"],
            "Throughput_1_10": "",
            "Quality_1_10": "",
            "Focus_1_10": "",
            "Stability_1_10": "",
            "Overall_1_10": blankish(rev.get("Overall_1_10")),
            "Wins": wins,
            "Misses": blankish(rev.get("Misses")),
            "Root_cause": "",
            "Keep": blankish(rev.get("Keep")),
            "Stop": blankish(rev.get("Stop")),
            "Start": blankish(rev.get("Start")),
            "Energy": blankish(rev.get("Energy")),
            "Next_priority": blankish(rev.get("Next_priority")),
            "Reviewed_on": "",
            "Note": "Từ DAILY Metrics review" if rev else "",
        }

    # rebuild groups for reopen
    by_week: dict[str, list] = defaultdict(list)
    by_month: dict[str, list] = defaultdict(list)
    by_q: dict[str, list] = defaultdict(list)
    year_rows: list[dict] = []
    for r in rows:
        d = r.get("date", "")
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            continue
        if dt.year != year:
            continue
        year_rows.append(r)
        by_week[r.get("week", "")].append(r)
        by_month[r.get("month", "")].append(r)
        q = (dt.month - 1) // 3 + 1
        by_q[f"{year}-Q{q}"].append(r)

    w_cols = perf_columns("Week", "Tuần ISO — 1 hàng review / tuần")
    w_recs = [
        _perf_rec("Week", x["Week"], x, by_week.get(x["Week"], []))
        for x in agg["weekly"]
    ]
    _write_sheet_rows(wb.create_sheet("Perf_Weekly"), w_cols, w_recs, [c[0] for c in w_cols])

    m_cols = perf_columns("Month", "Tháng YYYY-MM — review cuối tháng")
    m_recs = [
        _perf_rec("Month", x["Month"], x, by_month.get(x["Month"], []))
        for x in agg["monthly"]
    ]
    _write_sheet_rows(wb.create_sheet("Perf_Monthly"), m_cols, m_recs, [c[0] for c in m_cols])

    qperf_cols = perf_columns("Quarter", "Quý YYYY-Qn")
    qperf_recs = [
        _perf_rec("Quarter", x["Quarter"], x, by_q.get(x["Quarter"], []))
        for x in agg["quarterly"]
    ]
    _write_sheet_rows(
        wb.create_sheet("Perf_Quarterly"), qperf_cols, qperf_recs, [c[0] for c in qperf_cols]
    )

    yperf_cols = perf_columns("Year", "Năm")
    yperf_recs = [
        _perf_rec("Year", x["Year"], x, year_rows) for x in agg["yearly"]
    ]
    _write_sheet_rows(
        wb.create_sheet("Perf_Yearly"), yperf_cols, yperf_recs, [c[0] for c in yperf_cols]
    )

    # Rubric sheet
    rub = wb.create_sheet("Perf_Rubric")
    rub["A1"] = "Rubric chấm performance (Farming / work tracking)"
    rub["A1"].font = Font(bold=True, size=14)
    rub["A2"] = (
        "Điền Perf_Weekly / Monthly / Quarterly / Yearly. "
        "Cột snapshot (Items…Blockers) lấy từ WORK_METRICS — "
        "cột điểm và Keep/Stop/Start do bạn điền khi review. "
        "Không dùng làm điểm nhân sự; chỉ input tự phản tư."
    )
    rub["A2"].alignment = Alignment(wrap_text=True)
    rub.merge_cells("A2:E2")
    rub.row_dimensions[2].height = 40
    for i, h in enumerate(["Trục", "Ý nghĩa", "1–3 yếu", "4–7 ổn", "8–10 mạnh"], start=1):
        rub.cell(4, i, h)
        rub.cell(4, i).font = Font(bold=True, color="FFFFFF")
        rub.cell(4, i).fill = PatternFill("solid", fgColor="922B21")
    for i, row in enumerate(PERF_RUBRIC_ROWS, start=5):
        for c, val in enumerate(row, start=1):
            rub.cell(i, c, val)
            rub.cell(i, c).border = THIN
            rub.cell(i, c).alignment = Alignment(wrap_text=True, vertical="top")
    rub.column_dimensions["A"].width = 14
    rub.column_dimensions["B"].width = 28
    rub.column_dimensions["C"].width = 28
    rub.column_dimensions["D"].width = 32
    rub.column_dimensions["E"].width = 28

    # Projects
    proj_cols = [
        ("Project_ID", "Mã"),
        ("Name", "Tên hệ / dự án"),
        ("Status", "Active/Paused/… map từ STATUS"),
        ("Priority", "P0–P3"),
        ("Start", "Ngày bắt đầu nếu biết"),
        ("End", "Deadline"),
        ("Goal_ID", "Gắn goal"),
        ("Progress_pct", "Để trống hoặc tự điền"),
        ("Next_action", "Nợ chính từ STATUS"),
        ("Note", "Ghi chú"),
    ]
    _write_sheet_rows(wb.create_sheet("Projects"), proj_cols, projects, [c[0] for c in proj_cols])

    # Goals
    goal_cols = [
        ("Goal_ID", "Mã"),
        ("Title", "Objective"),
        ("Horizon", "Year/Quarter/Month"),
        ("Cycle", "Kỳ"),
        ("Status", "Active/…"),
        ("Why", "Lý do"),
        ("Progress_pct", "%"),
        ("Review_note", "Ghi chú"),
    ]
    goals = [
        {
            "Goal_ID": GOAL_ID,
            "Title": "Farming — combat presentation / Session Timeline usable cho designer",
            "Horizon": "Year",
            "Cycle": str(year),
            "Status": "Active",
            "Why": "Ship feel combat + authoring timeline; Docs DAILY/METRICS đo throughput",
            "Progress_pct": "",
            "Review_note": "Wave 1–3 (002–007) shipped; còn playtest + story 001",
        }
    ]
    _write_sheet_rows(wb.create_sheet("Goals"), goal_cols, goals, [c[0] for c in goal_cols])

    # Timeline
    tl_cols = [
        ("Milestone_ID", "Mã"),
        ("Project_ID", "Dự án"),
        ("Goal_ID", "Goal"),
        ("Title", "Mốc / Open item"),
        ("Due_date", "Hạn"),
        ("Status", "Todo/Doing/Done"),
        ("Deliverable", "Kết quả"),
        ("Note", "Nguồn"),
    ]
    _write_sheet_rows(wb.create_sheet("Timeline"), tl_cols, timeline, [c[0] for c in tl_cols])

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Farming WORK_METRICS → Excel work tracking")
    parser.add_argument(
        "--farming-docs",
        type=Path,
        default=DEFAULT_FARMING_DOCS,
        help="Thư mục Docs của game Farming",
    )
    parser.add_argument("--year", type=int, default=date.today().year)
    parser.add_argument("--out", type=Path, default=None)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    docs = args.farming_docs.resolve()
    metrics_path = docs / "WORK_METRICS.md"
    if not metrics_path.exists():
        raise SystemExit(f"Không thấy {metrics_path}")

    rows = parse_metrics_table(metrics_path)
    projects = parse_status_projects(docs / "STATUS.md")
    timeline = parse_epic_timeline(
        docs / "production" / "epics" / "presentation-sequence" / "EPIC.md"
    ) + parse_backlog_timeline(docs / "BACKLOG.md")

    out = args.out or (OUT_DIR / f"work-tracking-farming-{args.year}.xlsx")
    out = out.resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists() and not args.force:
        raise SystemExit(f"Đã có {out} — dùng --force để ghi đè.")

    wb = build_workbook(rows, args.year, projects, timeline, docs)
    polish_workbook(wb)
    wb.save(out)
    print(f"Đã import {len(rows)} hàng metrics → {out}")
    print(f"Projects: {len(projects)} · Timeline: {len(timeline)}")
    print("Sheets: … · Perf_Weekly/Monthly/Quarterly/Yearly · Perf_Rubric")
    print("Format: tab màu · zebra · status/effort/type/outcome tô màu")
    print("Review: Metrics review (DAILY) đổ vào Perf_* nếu có; còn lại điền tay")


if __name__ == "__main__":
    main()
