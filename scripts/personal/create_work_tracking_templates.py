#!/usr/bin/env python3
"""Tạo workbook Excel: dự án, mục tiêu (OKR), timeline, productivity.

Chu kỳ: ngày · tuần · tháng · quý · năm.
Khớp tinh thần:
  - templates/okr-planning.md
  - templates/weekly-review.md
  - templates/productivity/monthly-review-template.md
  - templates/productivity/learning-project-canvas.md

Usage:
  python scripts/personal/create_work_tracking_templates.py
  python scripts/personal/create_work_tracking_templates.py --year 2026
  python scripts/personal/create_work_tracking_templates.py --year 2026 --force
"""

from __future__ import annotations

import argparse
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_tracking_style import polish_workbook  # noqa: E402
from perf_review_schema import PERF_RUBRIC_ROWS, perf_columns  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "data" / "personal"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MEANING_FILL = PatternFill("solid", fgColor="FFF2CC")
MEANING_FONT = Font(italic=True, size=9, color="595959")
SECTION_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

HEADER_ROW = 1
MEANING_ROW = 2
DATA_START = 3

STATUS_PROJECT = '"Idea,Active,Paused,Done,Dropped"'
STATUS_GOAL = '"Draft,Active,Done,Dropped"'
STATUS_KR = '"Not started,On track,At risk,Done,Missed"'
STATUS_MILESTONE = '"Todo,Doing,Done,Skipped"'
PRIORITY = '"P0,P1,P2,P3"'
HORIZON = '"Year,Quarter,Month"'

SHEET_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "Projects": [
        ("Project_ID", "Mã dự án ngắn, ví dụ P-docs, P-game — dùng để nối Timeline"),
        ("Name", "Tên dự án dễ nhớ"),
        ("Status", "Idea / Active / Paused / Done / Dropped"),
        ("Priority", "P0 gấp · P1 quan trọng · P2–P3 thấp hơn"),
        ("Start", "Ngày bắt đầu (yyyy-mm-dd)"),
        ("End", "Deadline / ngày kết thúc dự kiến"),
        ("Goal_ID", "Mã mục tiêu liên quan (sheet Goals), nếu có"),
        ("Progress_pct", "Tiến độ 0–100"),
        ("Next_action", "Việc nhỏ tiếp theo (concrete, làm được hôm nay/mai)"),
        ("Note", "Ghi chú ngắn"),
    ],
    "Goals": [
        ("Goal_ID", "Mã mục tiêu, ví dụ G-2026-career, O1"),
        ("Title", "Mục tiêu định tính — ngắn, có cảm hứng (Objective)"),
        ("Horizon", "Year / Quarter / Month — chu kỳ chính của mục tiêu"),
        ("Cycle", "Kỳ cụ thể: 2026 · 2026-Q3 · 2026-08"),
        ("Status", "Draft / Active / Done / Dropped"),
        ("Why", "Vì sao quan trọng với bạn (động lực)"),
        ("Progress_pct", "Tiến độ tổng 0–100 (rollup từ KR nếu muốn)"),
        ("Review_note", "Ghi chú khi review cuối kỳ"),
    ],
    "Key_Results": [
        ("KR_ID", "Mã kết quả then chốt, ví dụ KR1, KR-G1-1"),
        ("Goal_ID", "Thuộc Goal/Objective nào"),
        ("Title", "KR đo được — ví dụ: ship 3 module / đọc 4 sách"),
        ("Metric_start", "Số bắt đầu (baseline)"),
        ("Metric_target", "Số đích cần đạt"),
        ("Metric_current", "Số hiện tại — cập nhật khi review"),
        ("Unit", "Đơn vị: giờ, bài, %, người…"),
        ("Progress_pct", "0–100; có thể = current/target"),
        ("Status", "Not started / On track / At risk / Done / Missed"),
    ],
    "Timeline": [
        ("Milestone_ID", "Mã mốc, ví dụ M1, M-launch"),
        ("Project_ID", "Thuộc dự án nào (sheet Projects)"),
        ("Goal_ID", "Hoặc gắn mục tiêu (nếu không phải project)"),
        ("Title", "Tên mốc / deliverable"),
        ("Due_date", "Hạn hoàn thành"),
        ("Status", "Todo / Doing / Done / Skipped"),
        ("Deliverable", "Sản phẩm cụ thể khi xong (file, demo, PR…)"),
        ("Note", "Phụ thuộc / rủi ro"),
    ],
    "Prod_Daily": [
        ("Date", "Ngày"),
        ("Deep_work_h", "Giờ deep work thật (không tính họp/mail loãng)"),
        ("Focus_blocks", "Số block tập trung hoàn thành (thường 1–3)"),
        ("Tasks_planned", "Số việc đã lên kế hoạch trong ngày"),
        ("Tasks_done", "Số việc hoàn thành"),
        ("Learning_h", "Giờ học / deliberate practice"),
        ("Project_ID", "Dự án chính hôm nay (nếu có)"),
        ("Energy_work_1_10", "Năng lượng lúc làm việc · 1 kiệt · 10 đầy"),
        ("One_win", "1 thắng lợi nhỏ trong ngày"),
        ("Blocker", "Thứ chặn tiến độ (nếu có)"),
    ],
    "Prod_Weekly": [
        ("Week", "Tuần ISO: YYYY-Www, ví dụ 2026-W33"),
        ("Week_start", "Ngày thứ Hai của tuần"),
        ("Deep_work_h", "Tổng giờ deep work trong tuần"),
        ("Tasks_done", "Số việc hoàn thành (ước lượng OK)"),
        ("Learning_h", "Tổng giờ học trong tuần"),
        ("Big_wins", "1–3 thành công đáng nhớ"),
        ("Challenges", "Khó khăn / miss chính"),
        ("Lesson", "Bài học mang sang tuần sau"),
        ("Plan_next", "Ưu tiên tuần tới (≤3)"),
        ("Score_1_10", "Tự chấm hiệu suất tuần · 1 tệ · 10 rất tốt"),
    ],
    "Prod_Monthly": [
        ("Month", "Tháng YYYY-MM"),
        ("Theme", "Chủ đề / focus tháng (1 câu)"),
        ("Deep_work_h", "Tổng deep work tháng (ước)"),
        ("Wins", "3 big wins"),
        ("Challenges", "Thách thức + nguyên nhân ngắn"),
        ("Energy_givers", "Việc/người/thói quen cho năng lượng"),
        ("Energy_drainers", "Việc hút năng lượng — cắt hoặc giới hạn"),
        ("OKR_check", "OKR/Goal tháng này: on track? ghi nhanh"),
        ("Score_1_10", "Tự chấm hiệu suất tháng"),
    ],
    "Prod_Quarterly": [
        ("Quarter", "Quý: YYYY-Q1 … YYYY-Q4"),
        ("Theme", "Chủ đề quý"),
        ("Objectives", "Tóm tắt Objective đang chase"),
        ("KR_hit_pct", "% KR đạt / gần đạt (0–100)"),
        ("Wins", "Thành tựu quý"),
        ("Misses", "KR/miss quan trọng + vì sao"),
        ("Pivot", "Giữ / bỏ / đổi gì quý sau"),
        ("Score_1_10", "Tự chấm hiệu suất quý"),
    ],
    "Prod_Yearly": [
        ("Year", "Năm"),
        ("Theme", "Một câu định hướng cả năm"),
        ("Top_wins", "3–5 thắng lớn nhất năm"),
        ("Top_lessons", "Bài học đắt nhất"),
        ("Carry_forward", "Việc / mục tiêu mang sang năm sau"),
        ("Score_1_10", "Tự chấm cả năm"),
    ],
}

SHEET_COLUMNS["Perf_Weekly"] = perf_columns("Week", "Tuần ISO — 1 hàng review / tuần")
SHEET_COLUMNS["Perf_Monthly"] = perf_columns("Month", "Tháng YYYY-MM")
SHEET_COLUMNS["Perf_Quarterly"] = perf_columns("Quarter", "Quý YYYY-Qn")
SHEET_COLUMNS["Perf_Yearly"] = perf_columns("Year", "Năm")


def _style_header(ws, row: int = HEADER_ROW) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 30


def _style_meaning_row(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(MEANING_ROW, c)
        cell.fill = MEANING_FILL
        cell.font = MEANING_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN
    ws.row_dimensions[MEANING_ROW].height = 52


def _autosize(ws, min_w: int = 12, max_w: int = 30) -> None:
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        samples: list[str] = []
        for r in (HEADER_ROW, MEANING_ROW):
            v = ws.cell(r, idx).value
            if v:
                samples.append(str(v)[:60])
        width = max((len(s) for s in samples), default=min_w) + 2
        ws.column_dimensions[letter].width = max(min_w, min(width, max_w))


def _write_column_headers(ws, columns: list[tuple[str, str]]) -> int:
    for i, (name, meaning) in enumerate(columns, start=1):
        ws.cell(HEADER_ROW, i, name)
        ws.cell(MEANING_ROW, i, meaning)
        ws.cell(HEADER_ROW, i).comment = Comment(meaning, "Docs", width=300, height=90)
    _style_header(ws, HEADER_ROW)
    _style_meaning_row(ws, len(columns))
    return len(columns)


def _blank_rows(ws, nrows: int, ncols: int, date_cols: set[int] | None = None) -> int:
    """Tạo nrows dòng trống từ DATA_START. Trả về last row."""
    date_cols = date_cols or set()
    for i in range(nrows):
        r = DATA_START + i
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
            if c in date_cols:
                ws.cell(r, c).number_format = "yyyy-mm-dd"
    last = DATA_START + nrows - 1
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    return last


def _add_list_validation(ws, formula: str, ranges: list[str]) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    for ref in ranges:
        dv.add(ref)


def _add_pct_validation(ws, ranges: list[str]) -> None:
    dv = DataValidation(
        type="whole", operator="between", formula1="0", formula2="100", allow_blank=True
    )
    ws.add_data_validation(dv)
    for ref in ranges:
        dv.add(ref)


def _add_score_validation(ws, ranges: list[str]) -> None:
    dv = DataValidation(
        type="whole", operator="between", formula1="1", formula2="10", allow_blank=True
    )
    ws.add_data_validation(dv)
    for ref in ranges:
        dv.add(ref)


def _year_days(year: int) -> list[date]:
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    days: list[date] = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)
    return days


def _iso_weeks(year: int) -> list[tuple[str, date]]:
    """[(YYYY-Www, monday), ...] cho các tuần có ít nhất 1 ngày trong năm."""
    weeks: list[tuple[str, date]] = []
    seen: set[str] = set()
    for d in _year_days(year):
        iso_year, iso_week, _ = d.isocalendar()
        key = f"{iso_year:04d}-W{iso_week:02d}"
        if key in seen:
            continue
        seen.add(key)
        monday = d - timedelta(days=d.weekday())
        weeks.append((key, monday))
    return weeks


def _write_glossary(wb: Workbook) -> None:
    ws = wb.create_sheet("Cot_y_nghia", 0)
    ws["A1"] = "Chú thích cột — Work / Project / Goal / Productivity"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Dòng 2 màu vàng trên mỗi sheet = giải thích ngắn. "
        "Di chuột vào tên cột cũng hiện chú thích. "
        "Template gốc: templates/okr-planning.md · weekly-review.md · monthly-review-template.md"
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 40

    for i, h in enumerate(["Sheet", "Cột", "Ý nghĩa", "Gợi ý"], start=1):
        ws.cell(4, i, h)
    _style_header(ws, 4)

    row = 5
    for sheet, cols in SHEET_COLUMNS.items():
        for name, meaning in cols:
            tip = ""
            if "Status" in name or name == "Priority" or name == "Horizon":
                tip = "Chọn trong dropdown"
            elif "pct" in name.lower() or "KR_hit" in name:
                tip = "Số 0–100"
            elif "1_10" in name or "Score_1_10" in name:
                tip = "Số 1–10"
            ws.cell(row, 1, sheet)
            ws.cell(row, 2, name)
            ws.cell(row, 3, meaning)
            ws.cell(row, 4, tip)
            for c in range(1, 5):
                ws.cell(row, c).border = THIN
                ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:D{row - 1}"


def _write_legend(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Legend", 1)
    ws["A1"] = "Work tracking — bản đồ sheet"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Năm", str(year)),
        ("File", str(OUT_DIR / f"work-tracking-{year}.xlsx")),
        ("", ""),
        ("→ Đọc trước", "Sheet Cot_y_nghia"),
        ("", ""),
        ("Sheet", "Việc"),
        ("Projects", "Danh sách dự án + next action"),
        ("Goals", "Mục tiêu / Objective theo năm·quý·tháng"),
        ("Key_Results", "KR đo được gắn Goal_ID"),
        ("Timeline", "Mốc thời gian / deliverable"),
        ("Prod_Daily", f"Productivity từng ngày trong {year}"),
        ("Prod_Weekly", "Review từng tuần ISO"),
        ("Prod_Monthly", "Review 12 tháng"),
        ("Prod_Quarterly", "Review 4 quý"),
        ("Prod_Yearly", "Review cả năm"),
        ("Perf_Weekly…Yearly", "Đánh giá performance + Keep/Stop/Start"),
        ("Perf_Rubric", "Rubric thang điểm"),
        ("", ""),
        ("Cách dùng gợi ý", ""),
        ("1", "Điền Goals + Key_Results cho năm/quý"),
        ("2", "Tạo Projects + Timeline mốc"),
        ("3", "Mỗi tối: 1 dòng Prod_Daily"),
        ("4", "Chủ nhật: Prod_Weekly + Perf_Weekly"),
        ("5", "Cuối tháng/quý/năm: Perf_Monthly / Quarterly / Yearly"),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    tip = 3 + len(rows) + 1
    ws.cell(
        tip,
        1,
        "Health/habit vẫn ở tracking-YYYY-MM.xlsx (create_tracking_templates.py). "
        "File này chuyên dự án–mục tiêu–năng suất làm việc.",
    )
    ws.merge_cells(start_row=tip, start_column=1, end_row=tip, end_column=2)
    ws.cell(tip, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 64


def _write_projects(wb: Workbook) -> None:
    ws = wb.create_sheet("Projects")
    cols = SHEET_COLUMNS["Projects"]
    ncols = _write_column_headers(ws, cols)
    last = _blank_rows(ws, 40, ncols, date_cols={5, 6})
    _add_list_validation(ws, STATUS_PROJECT, [f"C{DATA_START}:C{last}"])
    _add_list_validation(ws, PRIORITY, [f"D{DATA_START}:D{last}"])
    _add_pct_validation(ws, [f"H{DATA_START}:H{last}"])


def _write_goals(wb: Workbook) -> None:
    ws = wb.create_sheet("Goals")
    cols = SHEET_COLUMNS["Goals"]
    ncols = _write_column_headers(ws, cols)
    last = _blank_rows(ws, 30, ncols)
    _add_list_validation(ws, HORIZON, [f"C{DATA_START}:C{last}"])
    _add_list_validation(ws, STATUS_GOAL, [f"E{DATA_START}:E{last}"])
    _add_pct_validation(ws, [f"G{DATA_START}:G{last}"])


def _write_key_results(wb: Workbook) -> None:
    ws = wb.create_sheet("Key_Results")
    cols = SHEET_COLUMNS["Key_Results"]
    ncols = _write_column_headers(ws, cols)
    last = _blank_rows(ws, 40, ncols)
    _add_list_validation(ws, STATUS_KR, [f"I{DATA_START}:I{last}"])
    _add_pct_validation(ws, [f"H{DATA_START}:H{last}"])


def _write_timeline(wb: Workbook) -> None:
    ws = wb.create_sheet("Timeline")
    cols = SHEET_COLUMNS["Timeline"]
    ncols = _write_column_headers(ws, cols)
    last = _blank_rows(ws, 50, ncols, date_cols={5})
    _add_list_validation(ws, STATUS_MILESTONE, [f"F{DATA_START}:F{last}"])


def _write_prod_daily(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Prod_Daily")
    cols = SHEET_COLUMNS["Prod_Daily"]
    ncols = _write_column_headers(ws, cols)
    days = _year_days(year)
    for i, d in enumerate(days):
        r = DATA_START + i
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
        ws.cell(r, 1, d.isoformat())
        ws.cell(r, 1).number_format = "yyyy-mm-dd"
    last = DATA_START + len(days) - 1
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    _add_score_validation(ws, [f"H{DATA_START}:H{last}"])


def _write_prod_weekly(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Prod_Weekly")
    cols = SHEET_COLUMNS["Prod_Weekly"]
    ncols = _write_column_headers(ws, cols)
    weeks = _iso_weeks(year)
    for i, (key, monday) in enumerate(weeks):
        r = DATA_START + i
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
        ws.cell(r, 1, key)
        ws.cell(r, 2, monday.isoformat())
        ws.cell(r, 2).number_format = "yyyy-mm-dd"
    last = DATA_START + len(weeks) - 1
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    _add_score_validation(ws, [f"J{DATA_START}:J{last}"])
    for letter in ("F", "G", "H", "I"):
        ws.column_dimensions[letter].width = 28


def _write_prod_monthly(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Prod_Monthly")
    cols = SHEET_COLUMNS["Prod_Monthly"]
    ncols = _write_column_headers(ws, cols)
    for m in range(1, 13):
        r = DATA_START + m - 1
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
        ws.cell(r, 1, f"{year:04d}-{m:02d}")
    last = DATA_START + 11
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    _add_score_validation(ws, [f"I{DATA_START}:I{last}"])
    for letter in ("B", "D", "E", "F", "G", "H"):
        ws.column_dimensions[letter].width = 28


def _write_prod_quarterly(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Prod_Quarterly")
    cols = SHEET_COLUMNS["Prod_Quarterly"]
    ncols = _write_column_headers(ws, cols)
    for q in range(1, 5):
        r = DATA_START + q - 1
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
        ws.cell(r, 1, f"{year:04d}-Q{q}")
    last = DATA_START + 3
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    _add_pct_validation(ws, [f"D{DATA_START}:D{last}"])
    _add_score_validation(ws, [f"H{DATA_START}:H{last}"])
    for letter in ("B", "C", "E", "F", "G"):
        ws.column_dimensions[letter].width = 30


def _write_prod_yearly(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Prod_Yearly")
    cols = SHEET_COLUMNS["Prod_Yearly"]
    ncols = _write_column_headers(ws, cols)
    # 1 hàng năm hiện tại + 2 hàng trống cho năm khác
    for i, y in enumerate([year, year + 1, year - 1]):
        r = DATA_START + i
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = THIN
        ws.cell(r, 1, y)
    last = DATA_START + 2
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
    _autosize(ws)
    _add_score_validation(ws, [f"F{DATA_START}:F{last}"])
    for letter in ("B", "C", "D", "E"):
        ws.column_dimensions[letter].width = 36
    note = last + 2
    ws.cell(note, 1, "Hàng Year-1 / Year+1 để so sánh hoặc draft trước — xóa nếu không cần.")
    ws.cell(note, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=ncols)


def _write_perf_rubric(wb: Workbook) -> None:
    rub = wb.create_sheet("Perf_Rubric")
    rub["A1"] = "Rubric chấm performance"
    rub["A1"].font = Font(bold=True, size=14)
    rub["A2"] = (
        "Dùng với Perf_Weekly / Monthly / Quarterly / Yearly. "
        "Không phải điểm nhân sự — chỉ tự phản tư."
    )
    rub["A2"].alignment = Alignment(wrap_text=True)
    rub.merge_cells("A2:E2")
    for i, h in enumerate(["Trục", "Ý nghĩa", "1–3 yếu", "4–7 ổn", "8–10 mạnh"], start=1):
        cell = rub.cell(4, i, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="922B21")
    for i, row in enumerate(PERF_RUBRIC_ROWS, start=5):
        for c, val in enumerate(row, start=1):
            rub.cell(i, c, val)
            rub.cell(i, c).border = THIN
            rub.cell(i, c).alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDE", (14, 28, 28, 32, 28)):
        rub.column_dimensions[col].width = w


def _write_perf_sheets(wb: Workbook, year: int) -> None:
    """Hàng kỳ trống + snapshot cột để điền điểm review."""
    weeks = _iso_weeks(year)

    def empty_perf(period_key: str, period: str) -> dict:
        return {
            period_key: period,
            "Items": "",
            "Done": "",
            "Shipped": "",
            "Ship_pct": "",
            "Est_hours": "",
            "Reopen_sum": "",
            "Blockers": "",
            "Throughput_1_10": "",
            "Quality_1_10": "",
            "Focus_1_10": "",
            "Stability_1_10": "",
            "Overall_1_10": "",
            "Wins": "",
            "Misses": "",
            "Root_cause": "",
            "Keep": "",
            "Stop": "",
            "Start": "",
            "Energy": "",
            "Next_priority": "",
            "Reviewed_on": "",
            "Note": "",
        }

    for sheet, key, periods in [
        ("Perf_Weekly", "Week", [w[0] for w in weeks]),
        ("Perf_Monthly", "Month", [f"{year:04d}-{m:02d}" for m in range(1, 13)]),
        ("Perf_Quarterly", "Quarter", [f"{year:04d}-Q{q}" for q in range(1, 5)]),
        ("Perf_Yearly", "Year", [str(year)]),
    ]:
        ws = wb.create_sheet(sheet)
        cols = SHEET_COLUMNS[sheet]
        ncols = _write_column_headers(ws, cols)
        recs = [empty_perf(key, p) for p in periods]
        for i, rec in enumerate(recs):
            r = DATA_START + i
            for c, col_name in enumerate([x[0] for x in cols], start=1):
                ws.cell(r, c, rec.get(col_name) or None)
                ws.cell(r, c).border = THIN
        last = DATA_START + len(recs) - 1
        ws.freeze_panes = "B3"
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncols)}{last}"
        _autosize(ws)
        _add_score_validation(
            ws,
            [
                f"I{DATA_START}:M{last}",  # Throughput…Overall approx cols 9-13
            ],
        )

    _write_perf_rubric(wb)


def create_workbook(year: int) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    _write_glossary(wb)
    _write_legend(wb, year)
    _write_projects(wb)
    _write_goals(wb)
    _write_key_results(wb)
    _write_timeline(wb)
    _write_prod_daily(wb, year)
    _write_prod_weekly(wb, year)
    _write_prod_monthly(wb, year)
    _write_prod_quarterly(wb, year)
    _write_prod_yearly(wb, year)
    _write_perf_sheets(wb, year)
    return wb


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Tạo Excel tracking dự án / mục tiêu / timeline / productivity"
    )
    parser.add_argument(
        "--year",
        type=int,
        default=date.today().year,
        help="Năm (mặc định: năm hiện tại)",
    )
    parser.add_argument("--out", type=Path, default=None, help="Đường dẫn file output")
    parser.add_argument("--force", action="store_true", help="Ghi đè nếu đã có file")
    args = parser.parse_args()

    year = args.year
    out = args.out or (OUT_DIR / f"work-tracking-{year}.xlsx")
    out = out.resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    if out.exists() and not args.force:
        raise SystemExit(f"Đã có {out} — thêm --force để ghi đè.")

    wb = create_workbook(year)
    polish_workbook(wb)
    wb.save(out)
    print(f"Đã tạo: {out}")
    print(
        "Sheets: Cot_y_nghia · Legend · Projects · Goals · Key_Results · Timeline · "
        "Prod_Daily · Prod_Weekly · Prod_Monthly · Prod_Quarterly · Prod_Yearly · "
        "Perf_* · Perf_Rubric"
    )
    print("Format: tab màu · zebra · Status/Priority tự tô khi chọn dropdown")
    print("Perf_*: chấm Throughput/Quality/Focus/Stability/Overall + Keep/Stop/Start")


if __name__ == "__main__":
    main()
